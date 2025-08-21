from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from flask_migrate import Migrate
from datetime import datetime, timedelta, date
import pytz
import os
import markdown
from sqlalchemy import func, desc, text
import re
from markupsafe import Markup
import json
from icalendar import Calendar, Event
from io import BytesIO
from flask_mail import Mail, Message

# Load environment variables from .env file
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv not installed, continue without it

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')

# Configure session to last for a very long time (1 year)
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=365)
app.config['SESSION_COOKIE_SECURE'] = False  # Set to True in production with HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Configure database URI - use instance directory for SQLite
DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL:
    # Convert postgres:// to postgresql:// for newer SQLAlchemy versions
    if DATABASE_URL.startswith('postgres://'):
        DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
    # Convert to use psycopg driver for psycopg3
    if DATABASE_URL.startswith('postgresql://'):
        DATABASE_URL = DATABASE_URL.replace('postgresql://', 'postgresql+psycopg://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
else:
    # For local development and Render with SQLite
    instance_path = os.path.join(app.instance_path, 'hubtracker.db')
    app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{instance_path}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Email configuration (optional - will be set from environment variables)
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'True').lower() == 'true'
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER')

# Import models and db
from models import db, User, Client, Membership, Project, Task, Log, UserProjectPin, UserTaskFlag, TIMEZONE, get_current_time, MembershipSupplement, Equipment, UserPreferences, ActivityLog, SchedulingSettings, EquipmentOperatingHours, EquipmentBlockedDate, EquipmentAppointment

# Initialize extensions
db.init_app(app)
migrate = Migrate(app, db)

# Initialize Flask-Mail
mail = Mail(app)

def send_appointment_notification(appointment):
    """Send email notification for new appointment using Flask-Mail"""
    try:
        # Check if email is configured
        if not app.config.get('MAIL_SERVER'):
            print("Warning: Email not configured, skipping notification")
            return
        
        # Create email content
        appointment_url = f"https://hubtracker.onrender.com/schedule/{appointment.id}"  # Update with your domain
        equipment_name = appointment.equipment.name
        user_name = appointment.user.full_name
        user_email = appointment.user.email
        
        # Convert times to Central Time for email display
        central = pytz.timezone('America/Chicago')
        start_time_central = appointment.start_time.astimezone(central)
        end_time_central = appointment.end_time.astimezone(central)
        
        start_time = start_time_central.strftime('%B %d, %Y at %I:%M %p')
        end_time = end_time_central.strftime('%I:%M %p')
        
        # Calculate duration in Central Time
        duration_hours = (end_time_central - start_time_central).total_seconds() / 3600
        duration = f"{duration_hours:.1f} hours"
        purpose = appointment.purpose or "Not specified"
        
        # Email subject (more personal, less automated)
        subject = f"Equipment Booking Confirmation - {equipment_name}"
        
        # Email body (HTML) - more personal, less automated
        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <h2 style="color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px;">
                    Equipment Booking Confirmation
                </h2>
                
                <p>Hello,</p>
                
                <p>A new equipment booking has been scheduled for the Neurotech Hub:</p>
                
                <div style="background-color: #f8f9fa; padding: 20px; border-radius: 8px; margin: 20px 0;">
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr>
                            <td style="padding: 8px 0; font-weight: bold; width: 120px;">Equipment:</td>
                            <td style="padding: 8px 0;">{equipment_name}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; font-weight: bold;">User:</td>
                            <td style="padding: 8px 0;">{user_name} ({user_email})</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; font-weight: bold;">Date & Time:</td>
                            <td style="padding: 8px 0;">{start_time} - {end_time}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; font-weight: bold;">Duration:</td>
                            <td style="padding: 8px 0;">{duration}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; font-weight: bold;">Purpose:</td>
                            <td style="padding: 8px 0;">{purpose}</td>
                        </tr>
                    </table>
                </div>
                
                <p><strong>View Details:</strong> <a href="{appointment_url}" style="color: #3498db;">{appointment_url}</a></p>
                
                <p>Best regards,<br>
                Neurotech Hub Team</p>
                
                <hr style="border: none; border-top: 1px solid #eee; margin: 30px 0;">
                <p style="color: #7f8c8d; font-size: 12px; text-align: center;">
                    This is an automated notification from the Hub Tracker scheduling system.<br>
                    Washington University in St. Louis
                </p>
            </div>
        </body>
        </html>
        """
        
        # Plain text version
        text_content = f"""
New Equipment Booking

Equipment: {equipment_name}
User: {user_name} ({user_email})
Date & Time: {start_time} - {end_time}
Duration: {duration}
Purpose: {purpose}

View Details: {appointment_url}

This is an automated notification from the Hub Tracker scheduling system.
        """
        
        # Create and send email
        msg = Message(
            subject=subject,
            recipients=[os.environ.get('CONTACT_EMAIL', 'neurotechhub@wustl.edu')],
            html=html_content,
            body=text_content,
            sender=app.config.get('MAIL_DEFAULT_SENDER', 'neurotechhub.notifications@gmail.com'),
            extra_headers={
                'X-Mailer': 'Hub Tracker Scheduling System',
                'X-Priority': '3',
                'X-MSMail-Priority': 'Normal',
                'Importance': 'Normal'
            }
        )
        
        mail.send(msg)
        print("Email notification sent successfully")
        
    except Exception as e:
        print(f"Error sending email notification: {e}")
        # Don't raise the exception - don't break appointment creation

# Template context processors
@app.context_processor
def inject_pinned_projects():
    def get_pinned_projects():
        if 'user_id' not in session:
            return []
        
        # Get pinned projects for current user
        pinned_project_ids = UserProjectPin.query.filter_by(
            user_id=session['user_id']
        ).all()
        
        project_ids = [pin.project_id for pin in pinned_project_ids]
        
        if not project_ids:
            return []
        
        # Get the actual projects with client info
        pinned_projects = Project.query.join(Client).filter(
            Project.id.in_(project_ids)
        ).order_by(Project.name.asc()).all()
        
        # Add open tasks count to each pinned project
        for project in pinned_projects:
            project.open_tasks_count = project.tasks.filter_by(is_complete=False).count()
        
        return pinned_projects
    
    def has_logged_today():
        if 'user_id' not in session:
            return True  # Don't show animation if not logged in
        
        today_start = get_current_time().replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = get_current_time().replace(hour=23, minute=59, second=59, microsecond=999999)
        
        has_logs = Log.query.filter(
            Log.user_id == session['user_id'],
            Log.created_at.between(today_start, today_end)
        ).first() is not None
        
        return has_logs
    
    def should_show_log_reminder():
        if 'user_id' not in session:
            return False  # Don't show reminder if not logged in
        
        current_time = get_current_time()
        
        # Only show reminder after 12PM (noon) Chicago time
        if current_time.hour < 12:
            return False
        
        # Check if user has logged today
        today_start = current_time.replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = current_time.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        has_logs = Log.query.filter(
            Log.user_id == session['user_id'],
            Log.created_at.between(today_start, today_end)
        ).first() is not None
        
        return not has_logs
    
    return dict(
        get_pinned_projects=get_pinned_projects,
        has_logged_today=has_logged_today(),
        should_show_log_reminder=should_show_log_reminder()
    )

# Template filters
@app.template_filter('markdown')
def markdown_filter(text):
    """Convert markdown text to HTML"""
    if not text:
        return ''
    return markdown.markdown(text, extensions=['nl2br', 'fenced_code'])

@app.template_filter('render_tags')
def render_tags(value):
    value = re.sub(r'@\[(.*?)\]', r'<span class="task-tag task-tag-user">@\1</span>', value)
    value = re.sub(r'#\[(.*?)\]', r'<span class="task-tag task-tag-project">#\1</span>', value)
    return Markup(value)

@app.template_filter('time_ago')
def time_ago(datetime_obj):
    """Convert datetime to 'time ago' format"""
    if not datetime_obj:
        return ''
    
    from datetime import datetime
    now = get_current_time()
    diff = now - datetime_obj
    
    if diff.days > 0:
        if diff.days == 1:
            return '1 day ago'
        else:
            return f'{diff.days} days ago'
    elif diff.seconds >= 3600:
        hours = diff.seconds // 3600
        if hours == 1:
            return '1 hour ago'
        else:
            return f'{hours} hours ago'
    elif diff.seconds >= 60:
        minutes = diff.seconds // 60
        if minutes == 1:
            return '1 minute ago'
        else:
            return f'{minutes} minutes ago'
    else:
        return 'just now'

@app.template_filter('currency')
def currency_filter(value):
    """Format number as currency with commas and 2 decimal places"""
    if value is None:
        return '$0.00'
    
    try:
        # Convert to float and format with commas and 2 decimal places
        formatted = "${:,.2f}".format(float(value))
        return formatted
    except (ValueError, TypeError):
        return '$0.00'

@app.template_filter('time_until')
def time_until(dt):
    """Calculate time until a datetime in Chicago timezone"""
    if not dt:
        return ''
    
    import pytz
    from datetime import datetime
    
    # Ensure both datetimes are timezone-aware and in Chicago time
    chicago_tz = pytz.timezone('America/Chicago')
    now = datetime.now(chicago_tz)
    
    # If dt is naive, assume it's in Chicago time
    if dt.tzinfo is None:
        dt = chicago_tz.localize(dt)
    # If dt is in a different timezone, convert to Chicago time
    elif dt.tzinfo != chicago_tz:
        dt = dt.astimezone(chicago_tz)
    
    time_diff = dt - now
    hours_until = time_diff.total_seconds() / 3600
    
    if hours_until < 0:
        hours = int(-hours_until)
        if hours < 1:
            minutes = int(-time_diff.total_seconds() / 60)
            return f"{minutes} minute{'s' if minutes != 1 else ''} ago"
        elif hours < 24:
            return f"{hours} hour{'s' if hours != 1 else ''} ago"
        else:
            days = int(hours / 24)
            return f"{days} day{'s' if days != 1 else ''} ago"
    elif hours_until < 1:
        minutes = int(hours_until * 60)
        return f"In {minutes} minute{'s' if minutes != 1 else ''}"
    elif hours_until < 24:
        hours = int(hours_until)
        return f"In {hours} hour{'s' if hours != 1 else ''}"
    elif hours_until < 48:
        return "Tomorrow"
    else:
        days = int(hours_until / 24)
        return f"In {days} day{'s' if days != 1 else ''}"

@app.template_filter('central_time')
def central_time(dt):
    """Convert a datetime to Central Time for display"""
    if not dt:
        return ''
    
    import pytz
    
    # If dt is naive, assume it's in UTC (from database)
    if dt.tzinfo is None:
        utc_tz = pytz.timezone('UTC')
        dt = utc_tz.localize(dt)
    
    # Convert to Central Time
    central_tz = pytz.timezone('America/Chicago')
    central_dt = dt.astimezone(central_tz)
    
    return central_dt

@app.template_filter('duration_hours_central')
def duration_hours_central(appointment):
    """Calculate duration in hours using Central Time"""
    if not appointment.start_time or not appointment.end_time:
        return 0
    
    import pytz
    
    # Convert both times to Central Time for calculation
    central_tz = pytz.timezone('America/Chicago')
    
    # If times are naive, assume they're in UTC
    start_time = appointment.start_time
    end_time = appointment.end_time
    
    if start_time.tzinfo is None:
        utc_tz = pytz.timezone('UTC')
        start_time = utc_tz.localize(start_time)
    if end_time.tzinfo is None:
        utc_tz = pytz.timezone('UTC')
        end_time = utc_tz.localize(end_time)
    
    # Convert to Central Time
    start_central = start_time.astimezone(central_tz)
    end_central = end_time.astimezone(central_tz)
    
    # Calculate duration
    delta = end_central - start_central
    hours = delta.total_seconds() / 3600
    
    # Debug output
    print(f"DEBUG duration_hours_central:")
    print(f"  Original start: {appointment.start_time}")
    print(f"  Original end: {appointment.end_time}")
    print(f"  Start Central: {start_central}")
    print(f"  End Central: {end_central}")
    print(f"  Duration hours: {hours}")
    
    return hours

# Add global functions to Jinja2 environment
app.jinja_env.globals.update(min=min)
app.jinja_env.filters['markdown'] = markdown_filter
app.jinja_env.filters['render_tags'] = render_tags
app.jinja_env.filters['time_ago'] = time_ago
app.jinja_env.filters['currency'] = currency_filter
app.jinja_env.filters['time_until'] = time_until
app.jinja_env.filters['central_time'] = central_time
app.jinja_env.filters['duration_hours_central'] = duration_hours_central

@app.route('/')
def index():
    # Public landing page - show general metrics without sensitive data
    from datetime import timedelta
    from sqlalchemy import func
    
    # Get public-friendly metrics
    total_projects = Project.query.filter_by(status='Active').count()  # Only active projects
    total_clients = Client.query.count()
    total_active_projects = Project.query.filter_by(status='Active').count()
    total_users = User.query.count()
    total_memberships = Membership.query.filter_by(status='Active').count()
    open_tasks = Task.query.filter_by(is_complete=False).count()
    total_equipment = Equipment.query.count()
    
    # Last 30 days activity (including today)
    thirty_days_ago = get_current_time() - timedelta(days=29)  # Changed from 30 to 29
    
    # Tasks completed in last 30 days (general count)
    tasks_completed_recently = Task.query.filter(
        Task.completed_on >= thirty_days_ago,
        Task.is_complete == True
    ).count()
    
    # Get daily task completion for chart (last 30 days, including today)
    activity_data = []
    for i in range(30):  # Back to 30 days
        date = thirty_days_ago + timedelta(days=i)
        date_start = date.replace(hour=0, minute=0, second=0, microsecond=0)
        date_end = date_start + timedelta(days=1)
        
        daily_tasks = Task.query.filter(
            Task.completed_on >= date_start,
            Task.completed_on < date_end,
            Task.is_complete == True
        ).count()
        
        activity_data.append({
            'date': date.strftime('%Y-%m-%d'),
            'tasks': daily_tasks
        })
    
    # Get trends data for chart (last 30 days) - Hours Logged and Projects Touched
    trends_data = []
    for i in range(30):
        date = thirty_days_ago + timedelta(days=i)
        date_start = date.replace(hour=0, minute=0, second=0, microsecond=0)
        date_end = date_start + timedelta(days=1)
        
        # Calculate total hours logged for this day (including touch logs as 0.5 hours each)
        detailed_hours = db.session.query(func.sum(Log.hours)).filter(
            Log.created_at >= date_start,
            Log.created_at < date_end,
            Log.is_touch.is_(False),
            Log.hours.isnot(None)
        ).scalar() or 0
        
        # Get actual hours from touch logs
        touch_hours = db.session.query(func.sum(Log.hours)).filter(
            Log.created_at >= date_start,
            Log.created_at < date_end,
            Log.is_touch.is_(True),
            Log.hours.isnot(None)
        ).scalar() or 0
        
        total_hours = detailed_hours + touch_hours
        
        # Calculate projects touched based on activity logs
        projects_touched = set()
        
        try:
            # Get activity logs for this day
            daily_activities = ActivityLog.query.filter(
                ActivityLog.created_at >= date_start,
                ActivityLog.created_at < date_end
            ).all()
            
            for activity in daily_activities:
                project_id = None
                
                if activity.activity_type == 'task_completed':
                    # Get project from task
                    if activity.new_value and 'project_id' in activity.new_value:
                        project_id = activity.new_value['project_id']
                
                elif activity.activity_type == 'project_status_change':
                    # Entity_id is the project_id
                    project_id = activity.entity_id
                
                elif activity.activity_type in ['time_logged', 'touch_logged']:
                    # Get project from log entry
                    if activity.entity_type == 'log':
                        log_entry = db.session.get(Log, activity.entity_id)
                        if log_entry:
                            project_id = log_entry.project_id
                
                if project_id:
                    projects_touched.add(project_id)
        
        except Exception as e:
            # If ActivityLog table doesn't exist or there's an error, default to 0
            projects_touched = set()
        
        trends_data.append({
            'date': date.strftime('%Y-%m-%d'),
            'hours': round(total_hours, 1),
            'projects_touched': len(projects_touched)
        })
    
    # Get task velocity data (tasks completed per day, last 30 days)
    task_velocity_data = []
    for i in range(30):
        date = thirty_days_ago + timedelta(days=i)
        date_start = date.replace(hour=0, minute=0, second=0, microsecond=0)
        date_end = date_start + timedelta(days=1)
        
        daily_tasks = Task.query.filter(
            Task.completed_on >= date_start,
            Task.completed_on < date_end,
            Task.is_complete == True
        ).count()
        
        task_velocity_data.append({
            'date': date.strftime('%Y-%m-%d'),
            'tasks_completed': daily_tasks
        })
    
    # Project Status Data for pie chart (exclude default project)
    project_status_data = {
        'Active': Project.query.filter(Project.status == 'Active', Project.is_default == False).count(),
        'Awaiting': Project.query.filter(Project.status == 'Awaiting', Project.is_default == False).count(),
        'Paused': Project.query.filter(Project.status == 'Paused', Project.is_default == False).count(),
        'Archived': Project.query.filter(Project.status == 'Archived', Project.is_default == False).count()
    }
    
    # Recent Activity for public display (limit 10 items, sanitized)
    public_activities = []
    try:
        recent_activities = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(10).all()
        
        for activity in recent_activities:
            activity_item = None
            
            if activity.activity_type == 'task_completed':
                activity_item = {
                    'type': 'task_completed',
                    'description': 'Task completed for a project',
                    'created_at': activity.created_at
                }
            elif activity.activity_type == 'project_status_change':
                old_status = activity.old_value.get('status') if activity.old_value else 'Unknown'
                new_status = activity.new_value.get('status') if activity.new_value else 'Unknown'
                activity_item = {
                    'type': 'project_status_change',
                    'description': f'A project status changed from {old_status} to {new_status}',
                    'created_at': activity.created_at
                }
            elif activity.activity_type == 'time_logged':
                activity_item = {
                    'type': 'time_logged',
                    'description': 'Time logged for a project',
                    'created_at': activity.created_at
                }
            elif activity.activity_type == 'touch_logged':
                activity_item = {
                    'type': 'touch_logged',
                    'description': 'Quick touch logged for a project',
                    'created_at': activity.created_at
                }
            elif activity.activity_type == 'project_created':
                activity_item = {
                    'type': 'project_created',
                    'description': 'New project created',
                    'created_at': activity.created_at
                }
            elif activity.activity_type == 'client_created':
                activity_item = {
                    'type': 'client_created',
                    'description': 'New client added',
                    'created_at': activity.created_at
                }
            elif activity.activity_type == 'membership_created':
                activity_item = {
                    'type': 'membership_created',
                    'description': 'New membership created',
                    'created_at': activity.created_at
                }
            elif activity.activity_type == 'user_created':
                activity_item = {
                    'type': 'user_created',
                    'description': 'New team member added',
                    'created_at': activity.created_at
                }
            
            if activity_item:
                public_activities.append(activity_item)
    
    except Exception as e:
        print(f"Warning: Error processing activities for landing page: {e}")
        public_activities = []
    
    # Update variable names for metrics partial compatibility
    total_members = total_memberships  # Rename for partial compatibility
    
    return render_template('landing.html',
                         total_members=total_members,
                         total_projects=total_projects,
                         total_clients=total_clients,
                         open_tasks=open_tasks,
                         activity_data=activity_data,
                         trends_data=trends_data,
                         task_velocity_data=task_velocity_data,
                         project_status_data=project_status_data,
                         public_activities=public_activities)

@app.route('/login', methods=['GET', 'POST'])
def login():
    # If user is already logged in, redirect to dashboard
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    # Check if this is first-time setup (no users exist)
    user_count = User.query.count()
    
    if user_count == 0:
        if request.method == 'POST':
            # Create first admin user
            first_name = request.form.get('first_name', '').strip()
            last_name = request.form.get('last_name', '').strip()
            email = request.form.get('email', '').strip()
            password = request.form.get('password', '').strip()
            
            if not first_name or not email or not password:
                flash('First name, email, and password are required', 'error')
                return render_template('login.html', first_time_setup=True)
            
            # Check if email already exists (shouldn't happen in first-time setup)
            existing_user = User.query.filter_by(email=email).first()
            if existing_user:
                flash('Email already exists', 'error')
                return render_template('login.html', first_time_setup=True)
            
            # Create admin user
            user = User(
                first_name=first_name,
                email=email,
                role='admin'
            )
            user.set_last_name(last_name)
            user.set_password(password)
            
            db.session.add(user)
            db.session.commit()
            
            # Log them in
            session['user_id'] = user.id
            session['user_name'] = user.full_name
            session['role'] = user.role
            session.permanent = True  # Make session permanent for extended lifetime
            flash('Welcome! Your admin account has been created.', 'success')
            return redirect(url_for('dashboard'))
        
        # Show first-time setup form
        return render_template('login.html', first_time_setup=True)
    
    # Normal login flow
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()
        
        if not email:
            flash('Email is required', 'error')
            return render_template('login.html')
        
        user = User.query.filter_by(email=email).first()
        if not user:
            flash('Invalid email or password', 'error')
            return render_template('login.html')
        
        # Only check password for admin users
        if user.role == 'admin':
            if not password:
                flash('Password is required for admin users', 'error')
                return render_template('login.html')
            if not user.check_password(password):
                flash('Invalid email or password', 'error')
                return render_template('login.html')
        
        # Only allow admin users to login for now
        if user.role != 'admin':
            flash('Only administrators can login at this time', 'error')
            return render_template('login.html')
        
        session['user_id'] = user.id
        session['user_name'] = user.full_name
        session['role'] = user.role
        session.permanent = True  # Make session permanent for extended lifetime
        return redirect(url_for('dashboard'))
    
    return render_template('login.html')

@app.route('/login/<int:user_id>')
def login_user(user_id):
    # This route is deprecated but kept for backward compatibility
    # Redirect to main login page
    return redirect(url_for('login'))

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = db.session.get(User, session['user_id'])
    
    # Get tasks assigned to current user - filter out completed (for dashboard widget)
    tasks_for_me_query = Task.query.filter(
        (Task.assigned_to == session['user_id']) &
        (Task.is_complete == False)
    ).order_by(Task.created_at.desc()).all()
    
    # Get tasks created by current user - filter out completed (for dashboard widget)
    tasks_i_created_query = Task.query.filter(
        (Task.created_by == session['user_id']) &
        (Task.is_complete == False)
    ).order_by(Task.created_at.desc()).all()
    
    # Serialize tasks for JSON
    tasks_for_me = [serialize_task(task) for task in tasks_for_me_query]
    tasks_i_created = [serialize_task(task) for task in tasks_i_created_query]
    
    # Get upcoming equipment appointments
    from datetime import datetime, timedelta
    import pytz
    
    # Use Central Time for display, but UTC for database queries
    central = pytz.timezone('America/Chicago')
    now_central = datetime.now(central)
    
    # Convert to UTC for database query (since database stores in UTC)
    utc = pytz.timezone('UTC')
    now_utc = now_central.astimezone(utc)
    today_start_utc = now_utc.replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Debug: Print query parameters
    print(f"\nDEBUG Dashboard Appointments Query:")
    print(f"Current time (Central): {now_central}")
    print(f"Current time (UTC): {now_utc}")
    print(f"Today start (UTC): {today_start_utc}")
    
    # First get all appointments to check what exists
    all_appointments = EquipmentAppointment.query.all()
    print(f"\nDEBUG All Appointments ({len(all_appointments)}):")
    for appt in all_appointments:
        print(f"ID: {appt.id}, Equipment: {appt.equipment_id}, "
              f"Start: {appt.start_time}, Status: {appt.status}")
    
    # Now run the filtered query - for admins, show all appointments including cancelled for today
    upcoming_appointments_query = EquipmentAppointment.query.filter(
        EquipmentAppointment.start_time >= today_start_utc  # Show all of today's appointments
    )
    
    # Only filter by user if not an admin
    if session.get('role') != 'admin':
        upcoming_appointments_query = upcoming_appointments_query.filter(
            EquipmentAppointment.user_id == session['user_id']
        )
    
    upcoming_appointments = upcoming_appointments_query.order_by(
        EquipmentAppointment.start_time.asc()
    ).limit(5).all()
    
    print(f"\nDEBUG Filtered Appointments ({len(upcoming_appointments)}):")
    for appt in upcoming_appointments:
        print(f"ID: {appt.id}, Start: {appt.start_time}, Status: {appt.status}")
    
    # Recent Activity (limit 10 items)
    recent_activities = []
    try:
        # Check if ActivityLog table exists by trying to query it
        recent_activities = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(10).all()
    except Exception as e:
        print(f"Warning: ActivityLog table not available: {e}")
        # Set empty list to prevent further errors
        recent_activities = []
    
    # Process activities for display - combine all into single list for proper ordering
    all_activities = []
    
    try:
        for activity in recent_activities:
            if activity.activity_type == 'task_completed':
                # Get the task details
                task = db.session.get(Task, activity.entity_id)
                if task:
                    all_activities.append({
                        'type': 'task_completed',
                        'data': serialize_task(task),
                        'created_at': activity.created_at
                    })
            elif activity.activity_type == 'task_created' and activity.entity_id:
                # Get the task details for assignments
                task = db.session.get(Task, activity.entity_id)
                if task:
                    all_activities.append({
                        'type': 'task_created',
                        'data': serialize_task(task),
                        'created_at': activity.created_at
                    })
            elif activity.activity_type == 'time_logged':
                # Get the log details
                log = db.session.get(Log, activity.entity_id)
                if log:
                    all_activities.append({
                        'type': 'time_logged',
                        'data': log,
                        'created_at': activity.created_at
                    })
            elif activity.activity_type == 'project_status_change':
                all_activities.append({
                    'type': 'project_status_change',
                    'data': activity,
                    'created_at': activity.created_at
                })
            elif activity.activity_type == 'client_created':
                all_activities.append({
                    'type': 'client_created',
                    'data': activity,
                    'created_at': activity.created_at
                })
            elif activity.activity_type == 'membership_supplement_added':
                all_activities.append({
                    'type': 'membership_supplement_added',
                    'data': activity,
                    'created_at': activity.created_at
                })
            elif activity.activity_type == 'user_created':
                all_activities.append({
                    'type': 'user_created',
                    'data': activity,
                    'created_at': activity.created_at
                })
            elif activity.activity_type == 'project_created':
                all_activities.append({
                    'type': 'project_created',
                    'data': activity,
                    'created_at': activity.created_at
                })
            elif activity.activity_type == 'membership_created':
                all_activities.append({
                    'type': 'membership_created',
                    'data': activity,
                    'created_at': activity.created_at
                })
    except Exception as e:
        print(f"Warning: Error processing activities: {e}")
    
    # Get active projects for kanban board (exclude default project)
    active_projects = Project.query.filter(Project.status == 'Active', Project.is_default == False).order_by(Project.name.asc()).all()
    awaiting_projects = Project.query.filter(Project.status == 'Awaiting', Project.is_default == False).order_by(Project.name.asc()).all()
    paused_projects = Project.query.filter(Project.status == 'Paused', Project.is_default == False).order_by(Project.name.asc()).all()
    
    # Get clients and users for project forms
    clients = Client.query.order_by(Client.name.asc()).all()
    users = User.query.filter_by(role='admin').order_by(User.first_name.asc()).all()
    
    # Get current day of week for greeting
    day_of_week = now_central.strftime('%A')
    
    return render_template('dashboard.html',
                         user=user,
                         tasks_for_me=tasks_for_me,
                         tasks_i_created=tasks_i_created,
                         all_activities=all_activities,
                         day_of_week=day_of_week,
                         upcoming_appointments=upcoming_appointments,
                         now=now_central,  # Pass the localized current time
                         # Kanban data
                         active_projects=active_projects,
                         awaiting_projects=awaiting_projects,
                         paused_projects=paused_projects,
                         # Form data
                         clients=clients,
                         users=users)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/api/projects')
def get_projects():
    if 'user_id' not in session:
        return {'projects': []}, 401
    
    # Order projects: default project first, then by most recently updated
    projects = Project.query.join(Client).filter(
        Project.status != 'Archived'
    ).order_by(
        Project.is_default.desc(),  # Default project first
        Project.updated_at.desc()   # Then most recently updated
    ).all()
    
    return {
        'projects': [
            {
                'id': p.id,
                'name': p.name,
                'client_name': p.client.name,
                'display_name': f"{p.name} ({p.client.name})",
                'is_default': p.is_default
            }
            for p in projects
        ]
    }

@app.route('/api/users')
def get_users():
    if 'user_id' not in session:
        return {'users': []}, 401
    
    users = User.query.all()
    return {
        'users': [
            {
                'id': u.id,
                'name': u.full_name,
                'first_name': u.first_name
            }
            for u in users
        ]
    }

@app.route('/api/memberships')
def get_memberships():
    if 'user_id' not in session:
        return {'memberships': []}, 401
    
    memberships = Membership.query.all()
    return {
        'memberships': [
            {
                'id': m.id,
                'title': m.title
            }
            for m in memberships
        ]
    }

@app.route('/api/clients')
def get_clients():
    if 'user_id' not in session:
        return {'clients': []}, 401
    
    clients = Client.query.all()
    return {
        'clients': [
            {
                'id': c.id,
                'name': c.name
            }
            for c in clients
        ]
    }

@app.route('/api/export-report', methods=['POST'])
def export_report():
    if 'user_id' not in session:
        return {'error': 'Not logged in'}, 401
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime, timedelta
        import io
        
        # Get form data
        start_date_str = request.json.get('start_date')
        end_date_str = request.json.get('end_date')
        filter_type = request.json.get('filter_type', 'all')
        filter_value = request.json.get('filter_value', '')
        
        # Parse dates
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create header sheet
        header_ws = wb.create_sheet("Report Info")
        
        # Add header information
        header_ws['A1'] = "Hub Tracker Report"
        header_ws['A1'].font = Font(size=16, bold=True)
        
        header_ws['A3'] = "Date Range:"
        header_ws['B3'] = f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}"
        
        header_ws['A4'] = "Filter:"
        if filter_type == 'all':
            header_ws['B4'] = "All Memberships and Clients"
        elif filter_type == 'membership':
            membership = db.session.get(Membership, filter_value)
            header_ws['B4'] = f"Membership: {membership.title if membership else 'Unknown'}"
        elif filter_type == 'client':
            client = db.session.get(Client, filter_value)
            header_ws['B4'] = f"Client: {client.name if client else 'Unknown'}"
        
        header_ws['A6'] = "Generated:"
        header_ws['B6'] = datetime.now().strftime('%B %d, %Y at %I:%M %p')
        
        # Auto-adjust column widths
        for column in header_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            header_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create data sheet
        data_ws = wb.create_sheet("Project Data")
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        membership_font = Font(bold=True, size=12)
        membership_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        client_font = Font(bold=True)
        client_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add headers
        headers = [
            "Membership", "Client", "Project", "Log Type", "User", "Date/Time", "Hours", "Cost", "Notes", "Status"
        ]
        for col, header in enumerate(headers, 1):
            cell = data_ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # Helper function to calculate project totals
        def calculate_project_totals(project_id, start_date, end_date):
            """Calculate total hours and cost for a project within date range"""
            from sqlalchemy import func
            from datetime import timedelta
            
            # Get detailed logs with actual hours and costs
            detailed_logs = db.session.query(
                func.sum(Log.hours).label('total_hours'),
                func.sum(Log.fixed_cost).label('total_cost')
            ).filter(
                Log.project_id == project_id,
                Log.created_at >= start_date,
                Log.created_at < end_date + timedelta(days=1),
                Log.is_touch.is_(False)
            ).first()
            
            # Get actual hours from touch logs
            touch_hours = db.session.query(func.sum(Log.hours)).filter(
                Log.project_id == project_id,
                Log.created_at >= start_date,
                Log.created_at < end_date + timedelta(days=1),
                Log.is_touch.is_(True),
                Log.hours.isnot(None)
            ).scalar() or 0
            
            total_hours = float(detailed_logs.total_hours or 0) + touch_hours
            total_cost = float(detailed_logs.total_cost or 0)
            
            return total_hours, total_cost
        
        # Helper function to calculate membership remaining
        def calculate_membership_remaining(membership_id):
            """Calculate remaining budget and time for a membership, using supplement-aware totals"""
            membership = db.session.get(Membership, membership_id)
            if not membership:
                return 0, 0
            # Use the supplement-aware properties directly
            total_budget = float(membership.total_budget or 0)
            total_time = float(membership.total_time or 0)
            # Calculate used cost and time for this membership (all clients/projects)
            clients = Client.query.filter_by(membership_id=membership_id).all()
            client_ids = [c.id for c in clients]
            projects = Project.query.filter(Project.client_id.in_(client_ids)).all()
            used_cost = 0
            used_time = 0
            for project in projects:
                hours, cost = calculate_project_totals(project.id, membership.start_date or datetime(2020, 1, 1), datetime.now())
                used_time += hours
                used_cost += cost
            remaining_budget = total_budget - used_cost
            remaining_time = total_time - used_time
            return remaining_budget, remaining_time
        
        # Get memberships based on filter
        if filter_type == 'membership':
            memberships = [db.session.get(Membership, filter_value)] if filter_value else []
        elif filter_type == 'client':
            client = db.session.get(Client, filter_value)
            memberships = [client.membership] if client and client.membership else []
        else:  # all
            memberships = Membership.query.all()
        
        # Build report data
        row = 2
        membership_summaries = []
        
        for membership in memberships:
            if not membership:
                continue
                
            # Add membership header (merged row for clarity)
            membership_cell = data_ws.cell(row=row, column=1, value=membership.title)
            membership_cell.font = membership_font
            membership_cell.fill = membership_fill
            membership_cell.border = border
            row += 1

            clients = Client.query.filter_by(membership_id=membership.id).all()
            membership_total_hours = 0
            membership_total_cost = 0

            for client in clients:
                projects = Project.query.filter_by(client_id=client.id).all()
                
                # Check if this client has any projects with activity in the date range
                client_has_activity = False
                for project in projects:
                    logs = Log.query.filter(
                        Log.project_id == project.id,
                        Log.created_at >= start_date,
                        Log.created_at < end_date + timedelta(days=1)
                    ).first()
                    if logs:
                        client_has_activity = True
                        break
                
                # Skip clients with no activity in the date range
                if not client_has_activity:
                    continue

                # Add client header (merged row for clarity)
                client_cell = data_ws.cell(row=row, column=2, value=client.name)
                client_cell.font = client_font
                client_cell.fill = client_fill
                client_cell.border = border
                row += 1

                client_total_hours = 0
                client_total_cost = 0

                for project in projects:
                    # Fetch all logs for this project in the date range first
                    logs = Log.query.filter(
                        Log.project_id == project.id,
                        Log.created_at >= start_date,
                        Log.created_at < end_date + timedelta(days=1)
                    ).order_by(Log.created_at).all()

                    # Skip projects with no activity in the date range
                    if not logs:
                        continue

                    # Add project header (merged row for clarity)
                    project_cell = data_ws.cell(row=row, column=3, value=project.name)
                    project_cell.font = Font(italic=True)
                    project_cell.border = border
                    row += 1

                    project_hours = 0
                    project_cost = 0

                    for log in logs:
                        log_type = "Project Touched" if log.is_touch else "Time/Cost Log"
                        user_name = log.user.full_name if log.user else "Unknown"
                        dt_str = log.created_at.strftime('%Y-%m-%d %I:%M %p') if log.created_at else ""
                        hours = float(log.hours or 0)
                        cost = float(log.fixed_cost or 0)
                        notes = log.notes or ""
                        status = project.status

                        # Write log row
                        data_ws.cell(row=row, column=1, value=membership.title).border = border
                        data_ws.cell(row=row, column=2, value=client.name).border = border
                        data_ws.cell(row=row, column=3, value=project.name).border = border
                        data_ws.cell(row=row, column=4, value=log_type).border = border
                        data_ws.cell(row=row, column=5, value=user_name).border = border
                        data_ws.cell(row=row, column=6, value=dt_str).border = border
                        data_ws.cell(row=row, column=7, value=hours).border = border
                        data_ws.cell(row=row, column=8, value=cost).border = border
                        data_ws.cell(row=row, column=9, value=notes).border = border
                        data_ws.cell(row=row, column=10, value=status).border = border

                        project_hours += hours if not log.is_touch else 0.5
                        project_cost += cost
                        client_total_hours += hours if not log.is_touch else 0.5
                        client_total_cost += cost
                        membership_total_hours += hours if not log.is_touch else 0.5
                        membership_total_cost += cost
                        row += 1

                    # Project summary row
                    if logs:
                        data_ws.cell(row=row, column=3, value=f"{project.name} - TOTAL")
                        data_ws.cell(row=row, column=7, value=round(project_hours, 2))
                        data_ws.cell(row=row, column=8, value=round(project_cost, 2))
                        data_ws.cell(row=row, column=10, value="PROJECT TOTAL")
                        for col in range(1, 11):
                            cell = data_ws.cell(row=row, column=col)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                            cell.border = border
                        row += 1

                # Client summary row
                if projects:
                    data_ws.cell(row=row, column=2, value=f"{client.name} - TOTAL")
                    data_ws.cell(row=row, column=7, value=round(client_total_hours, 2))
                    data_ws.cell(row=row, column=8, value=round(client_total_cost, 2))
                    data_ws.cell(row=row, column=10, value="CLIENT TOTAL")
                    for col in range(1, 11):
                        cell = data_ws.cell(row=row, column=col)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        cell.border = border
                    row += 1

            # Membership summary row
            if clients:
                data_ws.cell(row=row, column=1, value=f"{membership.title} - TOTAL")
                data_ws.cell(row=row, column=7, value=round(membership_total_hours, 2))
                data_ws.cell(row=row, column=8, value=round(membership_total_cost, 2))
                data_ws.cell(row=row, column=10, value="MEMBERSHIP TOTAL")
                for col in range(1, 11):
                    cell = data_ws.cell(row=row, column=col)
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    cell.border = border
                row += 1
                
                # Calculate and store membership remaining
                remaining_budget, remaining_time = calculate_membership_remaining(membership.id)
                membership_summaries.append({
                    'name': membership.title,
                    'total_hours': membership_total_hours,
                    'total_cost': membership_total_cost,
                    'remaining_budget': remaining_budget,
                    'remaining_time': remaining_time
                })
        
        # Unassociated Projects Section - only include when not filtering by specific membership or client
        if filter_type == 'all':
            unassoc_clients = Client.query.filter(Client.membership_id.is_(None)).all()
            if unassoc_clients:
                # Section header
                data_ws.cell(row=row, column=1, value="Unassociated Projects").font = Font(bold=True, size=12)
                data_ws.cell(row=row, column=1).fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                row += 1
            
            for client in unassoc_clients:
                projects = Project.query.filter_by(client_id=client.id).all()
                
                # Check if this client has any projects with activity in the date range
                client_has_activity = False
                for project in projects:
                    logs = Log.query.filter(
                        Log.project_id == project.id,
                        Log.created_at >= start_date,
                        Log.created_at < end_date + timedelta(days=1)
                    ).first()
                    if logs:
                        client_has_activity = True
                        break
                
                # Skip clients with no activity in the date range
                if not client_has_activity:
                    continue
                
                # Client header
                client_cell = data_ws.cell(row=row, column=2, value=client.name)
                client_cell.font = client_font
                client_cell.fill = client_fill
                client_cell.border = border
                row += 1
                
                client_total_hours = 0
                client_total_cost = 0
                
                for project in projects:
                    # Fetch all logs for this project in the date range first
                    logs = Log.query.filter(
                        Log.project_id == project.id,
                        Log.created_at >= start_date,
                        Log.created_at < end_date + timedelta(days=1)
                    ).order_by(Log.created_at).all()
                    
                    # Skip projects with no activity in the date range
                    if not logs:
                        continue
                    
                    # Project header
                    project_cell = data_ws.cell(row=row, column=3, value=project.name)
                    project_cell.font = Font(italic=True)
                    project_cell.border = border
                    row += 1
                    
                    project_hours = 0
                    project_cost = 0
                    
                    for log in logs:
                        log_type = "Project Touched" if log.is_touch else "Time/Cost Log"
                        user_name = log.user.full_name if log.user else "Unknown"
                        dt_str = log.created_at.strftime('%Y-%m-%d %I:%M %p') if log.created_at else ""
                        hours = float(log.hours or 0)
                        cost = float(log.fixed_cost or 0)
                        notes = log.notes or ""
                        status = project.status
                        
                        # Write log row
                        data_ws.cell(row=row, column=1, value="Unassociated").border = border
                        data_ws.cell(row=row, column=2, value=client.name).border = border
                        data_ws.cell(row=row, column=3, value=project.name).border = border
                        data_ws.cell(row=row, column=4, value=log_type).border = border
                        data_ws.cell(row=row, column=5, value=user_name).border = border
                        data_ws.cell(row=row, column=6, value=dt_str).border = border
                        data_ws.cell(row=row, column=7, value=hours).border = border
                        data_ws.cell(row=row, column=8, value=cost).border = border
                        data_ws.cell(row=row, column=9, value=notes).border = border
                        data_ws.cell(row=row, column=10, value=status).border = border
                        
                        project_hours += hours if not log.is_touch else 0.5
                        project_cost += cost
                        client_total_hours += hours if not log.is_touch else 0.5
                        client_total_cost += cost
                        row += 1
                    
                    # Project summary row
                    if logs:
                        data_ws.cell(row=row, column=3, value=f"{project.name} - TOTAL")
                        data_ws.cell(row=row, column=7, value=round(project_hours, 2))
                        data_ws.cell(row=row, column=8, value=round(project_cost, 2))
                        data_ws.cell(row=row, column=10, value="PROJECT TOTAL")
                        for col in range(1, 11):
                            cell = data_ws.cell(row=row, column=col)
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                            cell.border = border
                        row += 1
                
                # Client summary row
                if projects:
                    data_ws.cell(row=row, column=2, value=f"{client.name} - TOTAL")
                    data_ws.cell(row=row, column=7, value=round(client_total_hours, 2))
                    data_ws.cell(row=row, column=8, value=round(client_total_cost, 2))
                    data_ws.cell(row=row, column=10, value="CLIENT TOTAL")
                    for col in range(1, 11):
                        cell = data_ws.cell(row=row, column=col)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        cell.border = border
                    row += 1
        
        # Create summary sheet
        summary_ws = wb.create_sheet("Membership Summary")
        
        # Add summary headers
        summary_headers = ["Membership", "Total Hours (Period)", "Total Cost (Period)", "Remaining Budget", "Remaining Time"]
        for col, header in enumerate(summary_headers, 1):
            cell = summary_ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
        
        # Add summary data
        for i, summary in enumerate(membership_summaries, 2):
            summary_ws.cell(row=i, column=1, value=summary['name']).border = border
            summary_ws.cell(row=i, column=2, value=round(summary['total_hours'], 2)).border = border
            summary_ws.cell(row=i, column=3, value=round(summary['total_cost'], 2)).border = border
            summary_ws.cell(row=i, column=4, value=round(summary['remaining_budget'], 2)).border = border
            summary_ws.cell(row=i, column=5, value=round(summary['remaining_time'], 2)).border = border
        
        # Auto-adjust column widths for data sheet
        for column in data_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            data_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Auto-adjust column widths for summary sheet
        for column in summary_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            summary_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"hub_tracker_report_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Export error: {str(e)}")
        return {'error': f'Export failed: {str(e)}'}, 500

@app.route('/api/tasks')
def get_tasks():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    # Get all tasks with their related data
    tasks = Task.query.filter(
        Task.is_complete == False  # Only get incomplete tasks for main list
    ).join(
        User, User.id == Task.created_by, aliased=True
    ).outerjoin(
        Project, Project.id == Task.project_id
    ).outerjoin(
        Client, Client.id == Project.client_id
    ).outerjoin(
        User, User.id == Task.assigned_to, aliased=True
    ).outerjoin(
        User, User.id == Task.completed_by_user_id, aliased=True
    ).all()
    
    # Convert tasks to dictionary
    tasks_data = []
    for task in tasks:
        task_dict = {
            'id': task.id,
            'description': task.description,
            'is_complete': task.is_complete,
            'completed_on': task.completed_on.isoformat() if task.completed_on else None,
            'created_at': task.created_at.isoformat(),
            'project_id': task.project_id,
            'project_name': task.project.name if task.project else None,
            'client_name': task.project.client.name if task.project and task.project.client else None,
            'assigned_to_id': task.assigned_to,
            'assigned_to_name': task.assignee.full_name if task.assignee else None,
            'created_by_id': task.created_by,
            'created_by_name': task.creator.full_name if task.creator else None,
            'completed_by_id': task.completed_by_user_id,
            'completed_by_name': task.completer.full_name if task.completer else None,
            'is_flagged': bool(task.flags.filter_by(user_id=session['user_id']).first())
        }
        tasks_data.append(task_dict)
    
    return jsonify({'tasks': tasks_data})

@app.route('/api/completed-tasks')
def get_completed_tasks():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    # Get the 10 most recently completed tasks
    tasks = Task.query.filter(
        Task.is_complete == True
    ).outerjoin(
        Project, Project.id == Task.project_id
    ).outerjoin(
        Client, Client.id == Project.client_id
    ).order_by(Task.completed_on.desc()).limit(10).all()
    
    # Convert tasks to dictionary
    tasks_data = []
    for task in tasks:
        task_dict = {
            'id': task.id,
            'description': task.description,
            'is_complete': task.is_complete,
            'completed_on': task.completed_on.isoformat() if task.completed_on else None,
            'created_at': task.created_at.isoformat(),
            'project_id': task.project_id,
            'project_name': task.project.name if task.project else None,
            'client_name': task.project.client.name if task.project and task.project.client else None,
            'assigned_to_id': task.assigned_to,
            'assigned_to_name': task.assignee.full_name if task.assignee else None,
            'created_by_id': task.created_by,
            'created_by_name': task.creator.full_name if task.creator else None,
            'completed_by_id': task.completed_by_user_id,
            'completed_by_name': task.completer.full_name if task.completer else None,
            'is_flagged': bool(task.flags.filter_by(user_id=session['user_id']).first())
        }
        tasks_data.append(task_dict)
    
    return jsonify({'tasks': tasks_data})

@app.route('/add_task', methods=['POST'])
def add_task():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    description = request.form.get('description', '').strip()  # This contains the full text with tags
    project_id = request.form.get('project_id')
    assigned_to = request.form.get('assigned_to')
    
    # Debug: Print form data
    print(f"DEBUG add_task: description='{description}', project_id='{project_id}', assigned_to='{assigned_to}'")
    print(f"DEBUG add_task: All form data: {dict(request.form)}")
    
    if description:
        # If no project specified, use the most recently updated project
        if not project_id:
            default_project = Project.query.filter(
                Project.status != 'Archived'
            ).order_by(Project.updated_at.desc()).first()
            project_id = default_project.id if default_project else None
        
        task = Task(
            description=description,  # Store the full text with tags
            created_by=session['user_id'],
            project_id=int(project_id) if project_id else None,
            assigned_to=int(assigned_to) if assigned_to else None
        )
        db.session.add(task)
        db.session.commit()
        
        # Log task creation activity
        ActivityLog.log_activity(
            user_id=session['user_id'],
            activity_type='task_created',
            entity_type='task',
            entity_id=task.id,
            new_value={
                'description': task.description,
                'project_id': task.project_id,
                'assigned_to': task.assigned_to
            }
        )
    
    # Redirect back to the referring page or dashboard
    return redirect(request.referrer or url_for('dashboard'))

@app.route('/task/<int:task_id>/complete', methods=['POST'])
def complete_task(task_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    task = Task.query.get_or_404(task_id)
    was_complete = task.is_complete
    task.toggle_complete(session['user_id'])
    db.session.commit()
    
    # Log task completion activity (only if it was completed, not uncompleted)
    if not was_complete and task.is_complete:
        ActivityLog.log_activity(
            user_id=session['user_id'],
            activity_type='task_completed',
            entity_type='task',
            entity_id=task.id,
            new_value={
                'description': task.description,
                'project_id': task.project_id,
                'assigned_to': task.assigned_to
            }
        )
    
    return redirect(request.referrer or url_for('tasks'))

@app.route('/task/<int:task_id>/delete', methods=['POST'])
def delete_task(task_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    task = Task.query.get_or_404(task_id)
    db.session.delete(task)
    db.session.commit()
    
    return redirect(request.referrer or url_for('tasks'))

@app.route('/task/<int:task_id>/edit', methods=['POST'])
def edit_task(task_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    task = Task.query.get_or_404(task_id)
    description = request.form.get('description', '').strip()
    project_id = request.form.get('project_id')
    assigned_to = request.form.get('assigned_to')
    
    if description:
        task.description = description  # Store the full text with tags
        task.project_id = int(project_id) if project_id else None
        task.assigned_to = int(assigned_to) if assigned_to else None
        db.session.commit()
    
    return redirect(request.referrer or url_for('tasks'))

@app.route('/task/<int:task_id>')
def task_detail(task_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    task = Task.query.get_or_404(task_id)
    
    # Get logs for this task's project
    project_logs = []
    if task.project:
        project_logs = task.project.logs.order_by(Log.created_at.desc()).limit(5).all()
    
    # Check if current user has flagged this task
    is_flagged = UserTaskFlag.query.filter_by(
        user_id=session['user_id'], 
        task_id=task_id
    ).first() is not None
    
    # Get projects and users for edit form
    projects = Project.query.join(Client).order_by(Project.name.asc()).all()
    users = User.query.order_by(User.first_name.asc()).all()
    
    return render_template('task_detail.html', 
                         task=task,
                         project_logs=project_logs,
                         is_flagged=is_flagged,
                         projects=projects,
                         users=users)

@app.route('/debug/tasks')
def debug_tasks():
    """Debug route to check all tasks in database"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    tasks = Task.query.all()
    result = f"<h2>Debug: All Tasks in Database</h2>"
    result += f"<p>Total tasks: {len(tasks)}</p>"
    result += f"<p>Current user_id: {session['user_id']}</p>"
    
    for task in tasks:
        result += f"<div style='border: 1px solid #ccc; margin: 10px; padding: 10px;'>"
        result += f"<strong>Task {task.id}:</strong> {task.description}<br>"
        result += f"Created by: {task.created_by}<br>"
        result += f"Assigned to: {task.assigned_to}<br>"
        result += f"Project ID: {task.project_id}<br>"
        result += f"Complete: {task.is_complete}<br>"
        result += f"Created: {task.created_at}<br>"
        if task.creator:
            result += f"Creator name: {task.creator.full_name}<br>"
        if task.assignee:
            result += f"Assignee name: {task.assignee.full_name}<br>"
        result += f"</div>"
    
    return result

# Placeholder routes for navigation
def serialize_task(task):
    """Convert task object to dictionary with related data"""
    try:
        # Safely access project and client info
        project_name = None
        client_name = None
        if task.project:
            project_name = task.project.name
            if task.project.client:
                client_name = task.project.client.name
        
        # Safely access user relationships
        assigned_to_name = None
        if task.assignee:
            assigned_to_name = task.assignee.full_name
            
        creator_name = None
        if task.creator:
            creator_name = task.creator.full_name
            
        completed_by_name = None
        if task.completer:
            completed_by_name = task.completer.full_name
        
        # Check if task is flagged by current user
        is_flagged = False
        if 'user_id' in session:
            is_flagged = UserTaskFlag.query.filter_by(
                user_id=session['user_id'], 
                task_id=task.id
            ).first() is not None
        
        # Apply tag formatting to description for frontend display
        formatted_description = render_tags(task.description)
        
        return {
            'id': task.id,
            'description': task.description,
            'formatted_description': formatted_description,
            'original_input': task.description,  # Use description as original_input since it contains the full tagged text
            'is_complete': task.is_complete,
            'completed_on': task.completed_on.isoformat() if task.completed_on else None,
            'created_at': task.created_at.isoformat(),
            'project_id': task.project_id,
            'project_name': project_name,
            'client_name': client_name,
            'assigned_to_name': assigned_to_name,
            'assigned_to_id': task.assigned_to,
            'creator_name': creator_name,
            'completed_by_name': completed_by_name,
            'is_flagged': is_flagged,
        }
    except Exception as e:
        print(f"Error serializing task {task.id}: {e}")
        return {
            'id': task.id,
            'description': task.description,
            'formatted_description': task.description,
            'original_input': task.description,
            'is_complete': task.is_complete,
            'completed_on': None,
            'created_at': task.created_at.isoformat() if task.created_at else '',
            'project_id': task.project_id,
            'project_name': None,
            'client_name': None,
            'assigned_to_name': None,
            'assigned_to_id': None,
            'creator_name': None,
            'completed_by_name': None,
            'is_flagged': False,
        }

@app.route('/tasks')
def tasks():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    print(f"DEBUG: Current user_id: {session['user_id']}")
    
    # Get tasks assigned to current user - filter out completed
    tasks_for_me_query = Task.query.filter(
        (Task.assigned_to == session['user_id']) &
        (Task.is_complete == False)
    ).order_by(Task.created_at.desc()).all()
    
    # Get tasks created by current user - filter out completed
    tasks_i_created_query = Task.query.filter(
        (Task.created_by == session['user_id']) &
        (Task.is_complete == False)
    ).order_by(Task.created_at.desc()).all()
    
    print(f"DEBUG: Found {len(tasks_for_me_query)} tasks for me, {len(tasks_i_created_query)} tasks I created")
    
    # Get all active tasks ordered by project name
    all_tasks_query = Task.query.join(Project, Task.project_id == Project.id, isouter=True).filter(
        Task.is_complete == False
    ).order_by(Project.name.asc(), Task.created_at.desc()).all()
    print(f"DEBUG: Found {len(all_tasks_query)} total active tasks")
    
    # Get all completed tasks
    completed_tasks_query = Task.query.filter(
        Task.is_complete == True
    ).order_by(Task.completed_on.desc()).all()
    print(f"DEBUG: Found {len(completed_tasks_query)} completed tasks")
    
    # Serialize tasks with related data
    tasks_for_me = [serialize_task(task) for task in tasks_for_me_query]
    tasks_i_created = [serialize_task(task) for task in tasks_i_created_query]
    all_tasks = [serialize_task(task) for task in all_tasks_query]
    completed_tasks = [serialize_task(task) for task in completed_tasks_query]
    
    print(f"DEBUG: Serialized {len(tasks_for_me)} tasks for me, {len(tasks_i_created)} tasks I created, {len(all_tasks)} all_tasks, {len(completed_tasks)} completed_tasks")
    
    return render_template('tasks.html', 
                         tasks_for_me=tasks_for_me, 
                         tasks_i_created=tasks_i_created, 
                         all_tasks=all_tasks, 
                         completed_tasks=completed_tasks)

# PROJECTS ROUTES
@app.route('/projects')
def projects():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    from sqlalchemy import func
    
    # Get all projects and separate by status
    all_projects = Project.query.join(Client).order_by(Project.name.asc()).all()
    
    # Separate active/prospective from archived projects
    active_projects = []
    archived_projects = []
    
    for project in all_projects:
        # Add open tasks count and pinned status to each project
        project.open_tasks_count = project.tasks.filter_by(is_complete=False).count()
        project.is_pinned = UserProjectPin.query.filter_by(
            user_id=session['user_id'], 
            project_id=project.id
        ).first() is not None
        
        # Calculate used hours and budget from logs
        logs_summary = db.session.query(
            func.sum(Log.hours).label('total_hours'),
            func.sum(Log.fixed_cost).label('total_cost')
        ).filter_by(project_id=project.id).first()
        
        project.used_hours = float(logs_summary.total_hours) if logs_summary.total_hours else 0
        project.used_budget = float(logs_summary.total_cost) if logs_summary.total_cost else 0
        
        # Get membership budget info through client
        project.membership_time = None
        project.membership_budget = None
        project.hours_remaining = None
        project.budget_remaining = None
        
        if project.client and project.client.membership:
            membership = project.client.membership
            if membership.time:
                project.membership_time = membership.time
                project.hours_remaining = max(0, membership.time - project.used_hours)
            if membership.budget:
                project.membership_budget = membership.budget
                project.budget_remaining = max(0, membership.budget - project.used_budget)
        
        if project.status == 'Archived':
            archived_projects.append(project)
        else:
            active_projects.append(project)
    
    clients = Client.query.order_by(Client.name.asc()).all()
    users = User.query.filter_by(role='admin').order_by(User.first_name.asc()).all()
    
    return render_template('projects.html', 
                         active_projects=active_projects,
                         archived_projects=archived_projects,
                         clients=clients, 
                         users=users)

@app.route('/project/<int:project_id>')
def project_detail(project_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    from sqlalchemy import func
    
    project = Project.query.get_or_404(project_id)
    
    # Calculate used hours and budget from logs
    logs_summary = db.session.query(
        func.sum(Log.hours).label('total_hours'),
        func.sum(Log.fixed_cost).label('total_cost')
    ).filter_by(project_id=project.id).first()
    
    project.used_hours = float(logs_summary.total_hours) if logs_summary.total_hours else 0
    project.used_budget = float(logs_summary.total_cost) if logs_summary.total_cost else 0
    
    # Get membership budget info through client
    project.membership_time = None
    project.membership_budget = None
    project.hours_remaining = None
    project.budget_remaining = None
    
    if project.client and project.client.membership:
        membership = project.client.membership
        if membership.time:
            project.membership_time = membership.time
            project.hours_remaining = max(0, membership.time - project.used_hours)
        if membership.budget:
            project.membership_budget = membership.budget
            project.budget_remaining = max(0, membership.budget - project.used_budget)
    
    # Get open tasks for this project
    open_tasks = project.tasks.filter_by(is_complete=False).order_by(Task.created_at.desc()).all()
    
    # Get completed tasks for this project
    completed_tasks = project.tasks.filter_by(is_complete=True).order_by(Task.completed_on.desc()).all()
    
    # Get all logs for this project
    all_logs = project.logs.order_by(Log.created_at.desc()).all()
    
    # Check if current user has pinned this project
    is_pinned = UserProjectPin.query.filter_by(
        user_id=session['user_id'], 
        project_id=project_id
    ).first() is not None
    
    # Serialize tasks
    serialized_open_tasks = [serialize_task(task) for task in open_tasks]
    serialized_completed_tasks = [serialize_task(task) for task in completed_tasks]
    
    # Get clients and users for dropdowns
    clients = Client.query.order_by(Client.name.asc()).all()
    users = User.query.filter_by(role='admin').order_by(User.first_name.asc()).all()
    
    return render_template('project_detail.html', 
                         project=project, 
                         tasks=serialized_open_tasks,
                         completed_tasks=serialized_completed_tasks,
                         all_logs=all_logs,
                         is_pinned=is_pinned,
                         clients=clients,
                         users=users)

@app.route('/add_project', methods=['POST'])
def add_project():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    name = request.form.get('name', '').strip()
    client_id = request.form.get('client_id')
    project_lead_id = request.form.get('project_lead_id')
    notes = request.form.get('notes', '').strip()
    status = request.form.get('status', 'Active')
    is_default = request.form.get('is_default') == 'on'
    
    if not name:
        flash('Project name is required', 'error')
        return redirect(url_for('projects'))
    
    if not client_id:
        flash('Client is required', 'error')
        return redirect(url_for('projects'))
    
    # If setting as default, remove default from other projects
    if is_default:
        Project.query.filter_by(is_default=True).update({'is_default': False})
    
    project = Project(
        name=name,
        client_id=int(client_id),
        project_lead_id=int(project_lead_id) if project_lead_id else None,
        notes=notes,
        status=status,
        is_default=is_default
    )
    
    db.session.add(project)
    db.session.commit()
    
    # Log project creation activity
    ActivityLog.log_activity(
        user_id=session['user_id'],
        activity_type='project_created',
        entity_type='project',
        entity_id=project.id,
        new_value={
            'name': project.name,
            'client_id': project.client_id,
            'status': project.status,
            'project_lead_id': project.project_lead_id
        }
    )
    
    flash('Project created successfully', 'success')
    return redirect(url_for('projects'))

@app.route('/edit_project/<int:project_id>', methods=['POST'])
def edit_project(project_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    project = Project.query.get_or_404(project_id)
    
    name = request.form.get('name', '').strip()
    client_id = request.form.get('client_id')
    project_lead_id = request.form.get('project_lead_id')
    notes = request.form.get('notes', '').strip()
    status = request.form.get('status', 'Active')
    is_default = request.form.get('is_default') == 'on'
    
    if not name:
        flash('Project name is required', 'error')
        return redirect(url_for('projects'))
    
    if not client_id:
        flash('Client is required', 'error')
        return redirect(url_for('projects'))
    
    # If setting as default, remove default from other projects
    if is_default and not project.is_default:
        Project.query.filter_by(is_default=True).update({'is_default': False})
    
    project.name = name
    project.client_id = int(client_id)
    project.project_lead_id = int(project_lead_id) if project_lead_id else None
    project.notes = notes
    project.status = status
    project.is_default = is_default
    
    db.session.commit()
    flash('Project updated successfully', 'success')
    return redirect(url_for('projects'))

@app.route('/delete_project/<int:project_id>', methods=['POST'])
def delete_project(project_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Check if project exists
    project = db.session.get(Project, project_id)
    if not project:
        flash(f'Project with ID {project_id} not found', 'error')
        return redirect(url_for('projects'))
    
    # Check if project has tasks
    if project.tasks.count() > 0:
        flash('Cannot delete project with existing tasks', 'error')
        return redirect(url_for('projects'))
    
    db.session.delete(project)
    db.session.commit()
    flash('Project deleted successfully', 'success')
    return redirect(url_for('projects'))

# CLIENTS ROUTES
@app.route('/clients')
def clients():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    clients = Client.query.order_by(Client.name.asc()).all()
    memberships = Membership.query.all()
    users = User.query.filter_by(role='admin').order_by(User.first_name.asc()).all()
    
    return render_template('clients.html', clients=clients, memberships=memberships, users=users)

@app.route('/client/<int:client_id>')
def client_detail(client_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    client = Client.query.get_or_404(client_id)
    
    # Get all projects for this client
    projects = client.projects.order_by(Project.name.asc()).all()
    
    # Get all tasks for this client's projects
    tasks = Task.query.join(Project).filter(
        Project.client_id == client_id
    ).order_by(Task.created_at.desc()).all()
    
    # Serialize tasks
    serialized_tasks = [serialize_task(task) for task in tasks]
    
    # Get memberships for dropdown
    memberships = Membership.query.all()
    users = User.query.filter_by(role='admin').order_by(User.first_name.asc()).all()
    clients = Client.query.order_by(Client.name.asc()).all()
    
    return render_template('client_detail.html', 
                         client=client, 
                         projects=projects, 
                         tasks=serialized_tasks,
                         memberships=memberships,
                         users=users,
                         clients=clients)

@app.route('/add_client', methods=['POST'])
def add_client():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    name = request.form.get('name', '').strip()
    membership_id = request.form.get('membership_id')
    notes = request.form.get('notes', '').strip()
    
    if not name:
        flash('Client name is required', 'error')
        return redirect(url_for('clients'))
    
    client = Client(
        name=name,
        membership_id=int(membership_id) if membership_id else None,
        notes=notes
    )
    
    db.session.add(client)
    db.session.commit()
    
    # Log client creation activity
    ActivityLog.log_activity(
        user_id=session['user_id'],
        activity_type='client_created',
        entity_type='client',
        entity_id=client.id,
        new_value={
            'name': client.name,
            'membership_id': client.membership_id
        }
    )
    
    flash('Client created successfully', 'success')
    return redirect(url_for('clients'))

@app.route('/edit_client/<int:client_id>', methods=['POST'])
def edit_client(client_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    client = Client.query.get_or_404(client_id)
    
    name = request.form.get('name', '').strip()
    membership_id = request.form.get('membership_id')
    notes = request.form.get('notes', '').strip()
    
    if not name:
        flash('Client name is required', 'error')
        return redirect(url_for('clients'))
    
    client.name = name
    client.membership_id = int(membership_id) if membership_id else None
    client.notes = notes
    
    db.session.commit()
    flash('Client updated successfully', 'success')
    return redirect(url_for('clients'))

@app.route('/delete_client/<int:client_id>', methods=['POST'])
def delete_client(client_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    client = Client.query.get_or_404(client_id)
    
    # Check if client has projects
    if client.projects.count() > 0:
        flash('Cannot delete client with existing projects', 'error')
        return redirect(url_for('clients'))
    
    db.session.delete(client)
    db.session.commit()
    flash('Client deleted successfully', 'success')
    return redirect(url_for('clients'))

# MEMBERSHIPS ROUTES
@app.route('/memberships')
def memberships():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    memberships = Membership.query.all()
    
    return render_template('memberships.html', memberships=memberships)

@app.route('/membership/<int:membership_id>')
def membership_detail(membership_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    membership = Membership.query.get_or_404(membership_id)
    
    # Get all clients for this membership
    associated_clients = membership.clients.order_by(Client.name.asc()).all()
    
    # Get all available clients for the form
    all_clients = Client.query.order_by(Client.name.asc()).all()
    
    # Get all projects through associated clients
    projects = []
    for client in associated_clients:
        projects.extend(client.projects.all())
    
    # Sort projects by name
    projects.sort(key=lambda p: p.name)
    
    return render_template('membership_detail.html', 
                         membership=membership,
                         clients=associated_clients,
                         all_clients=all_clients,
                         projects=projects,
                         MembershipSupplement=MembershipSupplement)

@app.route('/add_membership', methods=['POST'])
def add_membership():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    title = request.form.get('title', '').strip()
    start_date_str = request.form.get('start_date')
    is_annual = request.form.get('is_annual') == 'on'
    cost_str = request.form.get('cost', '').strip()
    time_str = request.form.get('time', '').strip()
    budget_str = request.form.get('budget', '').strip()
    notes = request.form.get('notes', '').strip()
    
    if not title:
        flash('Title is required', 'error')
        return redirect(url_for('memberships'))
    
    start_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            start_date = TIMEZONE.localize(start_date)
        except ValueError:
            flash('Invalid start date format', 'error')
            return redirect(url_for('memberships'))
    
    cost = None
    if cost_str:
        try:
            cost = float(cost_str)
        except ValueError:
            flash('Invalid cost format', 'error')
            return redirect(url_for('memberships'))
    
    time = None
    if time_str:
        try:
            time = int(time_str)
        except ValueError:
            flash('Invalid time format', 'error')
            return redirect(url_for('memberships'))
    
    budget = None
    if budget_str:
        try:
            budget = float(budget_str)
        except ValueError:
            flash('Invalid budget format', 'error')
            return redirect(url_for('memberships'))
    
    membership = Membership(
        title=title,
        start_date=start_date,
        is_annual=is_annual,
        cost=cost,
        time=time,
        budget=budget,
        notes=notes
    )
    
    db.session.add(membership)
    db.session.commit()
    
    # Log membership creation activity
    ActivityLog.log_activity(
        user_id=session['user_id'],
        activity_type='membership_created',
        entity_type='membership',
        entity_id=membership.id,
        new_value={
            'title': membership.title,
            'cost': membership.cost,
            'time': membership.time,
            'budget': membership.budget,
            'is_annual': membership.is_annual
        }
    )
    
    flash('Membership created successfully', 'success')
    return redirect(url_for('memberships'))

@app.route('/edit_membership/<int:membership_id>', methods=['POST'])
def edit_membership(membership_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    membership = Membership.query.get_or_404(membership_id)
    
    title = request.form.get('title', '').strip()
    start_date_str = request.form.get('start_date')
    status = request.form.get('status', 'Active')
    is_annual = request.form.get('is_annual') == 'on'
    cost_str = request.form.get('cost', '').strip()
    time_str = request.form.get('time', '').strip()
    budget_str = request.form.get('budget', '').strip()
    notes = request.form.get('notes', '').strip()
    
    if not title:
        flash('Title is required', 'error')
        return redirect(url_for('memberships'))
    
    start_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            start_date = TIMEZONE.localize(start_date)
        except ValueError:
            flash('Invalid start date format', 'error')
            return redirect(url_for('memberships'))
    
    cost = None
    if cost_str:
        try:
            cost = float(cost_str)
        except ValueError:
            flash('Invalid cost format', 'error')
            return redirect(url_for('memberships'))
    
    time = None
    if time_str:
        try:
            time = int(time_str)
        except ValueError:
            flash('Invalid time format', 'error')
            return redirect(url_for('memberships'))
    
    budget = None
    if budget_str:
        try:
            budget = float(budget_str)
        except ValueError:
            flash('Invalid budget format', 'error')
            return redirect(url_for('memberships'))
    
    membership.title = title
    membership.start_date = start_date
    membership.status = status
    membership.is_annual = is_annual
    membership.cost = cost
    membership.time = time
    membership.budget = budget
    membership.notes = notes
    
    db.session.commit()
    flash('Membership updated successfully', 'success')
    
    # Redirect back to the membership detail page if we came from there
    referrer = request.headers.get('Referer', '')
    if f'/membership/{membership_id}' in referrer:
        return redirect(url_for('membership_detail', membership_id=membership_id))
    else:
        return redirect(url_for('memberships'))

@app.route('/membership/<int:membership_id>/update_clients', methods=['POST'])
def update_membership_clients(membership_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    membership = Membership.query.get_or_404(membership_id)
    
    # Get the list of client IDs that should be associated
    selected_client_ids = request.form.getlist('client_ids')
    
    # Convert to integers
    try:
        selected_client_ids = [int(cid) for cid in selected_client_ids]
    except ValueError:
        flash('Invalid client ID format', 'error')
        return redirect(url_for('membership_detail', membership_id=membership_id))
    
    # Get all clients that should be associated
    clients_to_associate = Client.query.filter(Client.id.in_(selected_client_ids)).all()
    
    # Get current associated clients
    current_clients = membership.clients.all()
    
    # Remove clients that are no longer selected
    for client in current_clients:
        if client.id not in selected_client_ids:
            client.membership_id = None
    
    # Add newly selected clients
    for client in clients_to_associate:
        if client not in current_clients:
            client.membership_id = membership.id
    
    db.session.commit()
    flash('Client associations updated successfully', 'success')
    return redirect(url_for('membership_detail', membership_id=membership_id))

@app.route('/delete_membership/<int:membership_id>', methods=['POST'])
def delete_membership(membership_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    membership = Membership.query.get_or_404(membership_id)
    
    # Check if membership has clients
    if membership.clients.count() > 0:
        flash('Cannot delete membership with existing clients', 'error')
        return redirect(url_for('memberships'))
    
    db.session.delete(membership)
    db.session.commit()
    flash('Membership deleted successfully', 'success')
    return redirect(url_for('memberships'))

@app.route('/kanban')
def kanban():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get all projects grouped by status
    active_projects = Project.query.filter_by(status='Active').order_by(Project.updated_at.desc()).all()
    awaiting_projects = Project.query.filter_by(status='Awaiting').order_by(Project.updated_at.desc()).all()
    paused_projects = Project.query.filter_by(status='Paused').order_by(Project.updated_at.desc()).all()
    archived_projects = Project.query.filter_by(status='Archived').order_by(Project.updated_at.desc()).all()
    
    # If no Awaiting/Paused projects exist yet, also check Prospective projects for Awaiting column
    if not awaiting_projects:
        awaiting_projects = Project.query.filter_by(status='Prospective').order_by(Project.updated_at.desc()).all()
    
    # Get clients and users for the add project modal
    clients = Client.query.order_by(Client.name.asc()).all()
    users = User.query.filter_by(role='admin').order_by(User.first_name.asc()).all()
    
    return render_template('kanban.html',
                         active_projects=active_projects,
                         awaiting_projects=awaiting_projects,
                         paused_projects=paused_projects,
                         archived_projects=archived_projects,
                         clients=clients,
                         users=users)

@app.route('/api/project/<int:project_id>/status', methods=['POST'])
def update_project_status(project_id):
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': 'Not authenticated'}), 401
    
    project = Project.query.get_or_404(project_id)
    data = request.get_json()
    new_status = data.get('status')
    
    # Validate status
    valid_statuses = ['Active', 'Awaiting', 'Paused', 'Archived']
    if new_status not in valid_statuses:
        return jsonify({'success': False, 'error': 'Invalid status'}), 400
    
    # Use the new update_status method that logs activity
    print(f"DEBUG: Updating project {project_id} status to {new_status}")
    project.update_status(new_status, session['user_id'])
    
    try:
        db.session.commit()
        return jsonify({'success': True, 'message': f'Project moved to {new_status}'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/analytics')
def analytics():
    """Analytics page (formerly reports)"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    from datetime import timedelta
    from sqlalchemy import func, desc
    import pytz
    
    # Last 30 days filter - get current time in Chicago timezone
    chicago_tz = pytz.timezone('America/Chicago')
    utc_tz = pytz.timezone('UTC')
    current_time_chicago = get_current_time()
    
    # Convert to UTC for database queries
    current_time_utc = current_time_chicago.astimezone(utc_tz)
    thirty_days_ago_utc = current_time_utc - timedelta(days=30)
    
    # Debug logging
    print(f"\n=== DEBUG: Analytics Timezone Info ===")
    print(f"Current time Chicago: {current_time_chicago}")
    print(f"Current time UTC: {current_time_utc}")
    print(f"Thirty days ago UTC: {thirty_days_ago_utc}")
    
    # Get basic stats for the top metrics
    total_members = Membership.query.filter_by(status='Active').count()
    total_projects = Project.query.count()
    total_clients = Client.query.count()
    open_tasks = Task.query.filter_by(is_complete=False).count()

    
    # Get task completion data for last 30 days (including today)
    twenty_nine_days_ago_utc = current_time_utc - timedelta(days=29)
    
    # Generate date series for last 30 days (including today)
    completion_data = []
    for i in range(30):
        # Calculate date in Chicago timezone for display
        # We want to go back from today (i=0) to 29 days ago (i=29)
        # So for i=0, we want today, for i=29, we want 29 days ago
        date_chicago = current_time_chicago - timedelta(days=i)
        date_start_chicago = date_chicago.replace(hour=0, minute=0, second=0, microsecond=0)
        date_end_chicago = date_start_chicago + timedelta(days=1)
        
        # Convert to UTC for database queries - handle timezone conversion properly
        if date_start_chicago.tzinfo is None:
            date_start_utc = chicago_tz.localize(date_start_chicago).astimezone(utc_tz)
        else:
            date_start_utc = date_start_chicago.astimezone(utc_tz)
            
        if date_end_chicago.tzinfo is None:
            date_end_utc = chicago_tz.localize(date_end_chicago).astimezone(utc_tz)
        else:
            date_end_utc = date_end_chicago.astimezone(utc_tz)
        
        # Debug logging for August 6th specifically
        debug_date = date_start_chicago.strftime('%Y-%m-%d')
        if debug_date == '2025-08-06':
            print(f"\n=== DEBUG: August 6th Hours Calculation ===")
            print(f"Date Chicago: {date_start_chicago} to {date_end_chicago}")
            print(f"Date UTC: {date_start_utc} to {date_end_utc}")
            print(f"Days from current: {i}")
            print(f"Current time Chicago: {current_time_chicago}")
            print(f"Calculated date Chicago: {date_chicago}")
        
        # All users completions for this day
        all_completions = Task.query.filter(
            Task.completed_on >= date_start_utc,
            Task.completed_on < date_end_utc,
            Task.is_complete.is_(True)
        ).count()
        
        # Current user completions for this day
        my_completions = Task.query.filter(
            Task.completed_on >= date_start_utc,
            Task.completed_on < date_end_utc,
            Task.completed_by_user_id == session['user_id'],
            Task.is_complete.is_(True)
        ).count()
        
        # All users hours for this day (including touch logs as 0.5 hours each)
        all_detailed_hours = db.session.query(func.sum(Log.hours)).filter(
            Log.created_at >= date_start_utc,
            Log.created_at < date_end_utc,
            Log.is_touch.is_(False),
            Log.hours.isnot(None)
        ).scalar() or 0
        
        # Get actual hours from touch logs
        all_touch_hours = db.session.query(func.sum(Log.hours)).filter(
            Log.created_at >= date_start_utc,
            Log.created_at < date_end_utc,
            Log.is_touch.is_(True),
            Log.hours.isnot(None)
        ).scalar() or 0
        all_hours = float(all_detailed_hours) + all_touch_hours
        
        # Current user hours for this day (including touch logs as 0.5 hours each)
        my_detailed_hours = db.session.query(func.sum(Log.hours)).filter(
            Log.created_at >= date_start_utc,
            Log.created_at < date_end_utc,
            Log.user_id == session['user_id'],
            Log.is_touch.is_(False),
            Log.hours.isnot(None)
        ).scalar() or 0
        
        # Get actual hours from touch logs for current user
        my_touch_hours = db.session.query(func.sum(Log.hours)).filter(
            Log.created_at >= date_start_utc,
            Log.created_at < date_end_utc,
            Log.user_id == session['user_id'],
            Log.is_touch.is_(True),
            Log.hours.isnot(None)
        ).scalar() or 0
        my_hours = float(my_detailed_hours) + my_touch_hours
        
        # Debug logging for August 6th specifically
        if debug_date == '2025-08-06':
            print(f"All detailed hours: {all_detailed_hours}")
            print(f"All touch hours: {all_touch_hours}")
            print(f"All hours total: {all_hours}")
            print(f"My detailed hours: {my_detailed_hours}")
            print(f"My touch hours: {my_touch_hours}")
            print(f"My hours total: {my_hours}")
            
            # Let's also check what logs exist for this user on this date
            user_logs = Log.query.filter(
                Log.created_at >= date_start_utc,
                Log.created_at < date_end_utc,
                Log.user_id == session['user_id']
            ).all()
            print(f"Found {len(user_logs)} logs for user {session['user_id']} on {debug_date}:")
            for log in user_logs:
                print(f"  - Log ID: {log.id}, Hours: {log.hours}, Created: {log.created_at}, Is Touch: {log.is_touch}")
                
            # Let's also check all logs for this date range to see what's in the database
            all_logs = Log.query.filter(
                Log.created_at >= date_start_utc,
                Log.created_at < date_end_utc
            ).all()
            print(f"Found {len(all_logs)} total logs on {debug_date}:")
            for log in all_logs:
                print(f"  - Log ID: {log.id}, User: {log.user_id}, Hours: {log.hours}, Created: {log.created_at}, Is Touch: {log.is_touch}")
        
        completion_data.append({
            'date': date_start_chicago.strftime('%Y-%m-%d'),
            'all_tasks': all_completions,
            'my_tasks': my_completions,
            'all_hours': round(all_hours, 1),
            'my_hours': round(my_hours, 1)
        })
    
    # Debug logging for the final completion_data
    print(f"\n=== DEBUG: Final Completion Data ===")
    print(f"Total entries: {len(completion_data)}")
    aug6_entry = next((entry for entry in completion_data if entry['date'] == '2025-08-06'), None)
    if aug6_entry:
        print(f"August 6th entry: {aug6_entry}")
    else:
        print("August 6th entry not found!")
        print("Available dates:", [entry['date'] for entry in completion_data])
    
    # Helper function to calculate total logged time for a project
    def calculate_project_logged_time(project_id, start_date):
        """Calculate total logged time for a project including actual touch log hours"""
        from sqlalchemy import func, case
        
        # Get detailed logs with actual hours
        detailed_hours = db.session.query(func.sum(Log.hours)).filter(
            Log.project_id == project_id,
            Log.created_at >= start_date,
            Log.is_touch.is_(False)
        ).scalar() or 0
        
        # Get actual hours from touch logs
        touch_hours = db.session.query(func.sum(Log.hours)).filter(
            Log.project_id == project_id,
            Log.created_at >= start_date,
            Log.is_touch.is_(True),
            Log.hours.isnot(None)
        ).scalar() or 0
        
        return float(detailed_hours) + touch_hours
    
    # Get most active projects by task completion (last 30 days)
    most_active_projects_raw = db.session.query(
        Project.name,
        Client.name.label('client_name'),
        func.count(Task.id).label('completed_tasks')
    ).join(Client).join(Task).filter(
        Task.is_complete.is_(True),
        Task.completed_on >= twenty_nine_days_ago_utc
    ).group_by(Project.id, Project.name, Client.name).order_by(desc('completed_tasks')).limit(8).all()
    
    # Convert to dictionaries for JSON serialization
    most_active_projects = []
    for project in most_active_projects_raw:
        most_active_projects.append({
            'name': project.name,
            'client_name': project.client_name,
            'completed_tasks': project.completed_tasks
        })
    
    # Get most active projects by logged time (last 30 days)
    most_logged_time_projects = []
    projects_with_logs = db.session.query(
        Project.id,
        Project.name,
        Client.name.label('client_name')
    ).join(Client).join(Log).filter(
        Log.created_at >= thirty_days_ago_utc
    ).group_by(Project.id, Project.name, Client.name).all()
    
    for project in projects_with_logs:
        total_hours = calculate_project_logged_time(project.id, thirty_days_ago_utc)
        if total_hours > 0:  # Only include projects with actual logged time
            most_logged_time_projects.append({
                'name': project.name,
                'client_name': project.client_name,
                'total_hours': total_hours
            })
    
    # Sort by total hours and limit to top 8
    most_logged_time_projects.sort(key=lambda x: x['total_hours'], reverse=True)
    most_logged_time_projects = most_logged_time_projects[:8]
    
    # LOG ANALYTICS (Last 30 Days)
    
    # Total log metrics
    total_logs_last_30 = Log.query.filter(Log.created_at >= thirty_days_ago_utc).count()
    touch_logs_last_30 = Log.query.filter(Log.created_at >= thirty_days_ago_utc, Log.is_touch.is_(True)).count()
    detailed_logs_last_30 = Log.query.filter(Log.created_at >= thirty_days_ago_utc, Log.is_touch.is_(False)).count()
    total_hours_last_30 = db.session.query(func.sum(Log.hours)).filter(Log.created_at >= thirty_days_ago_utc).scalar() or 0
    
    # Daily log activity for chart
    log_activity_data = []
    
    # If there are no logs at all, just return empty data for today
    if total_logs_last_30 == 0:
        today = get_current_time()
        log_activity_data.append({
            'date': today.strftime('%Y-%m-%d'),
            'touch_logs': 0,
            'detailed_logs': 0,
            'total_logs': 0
        })
    else:
        for i in range(30):
            # Calculate date in Chicago timezone for display
            # We want to go back from today (i=0) to 29 days ago (i=29)
            # So for i=0, we want today, for i=29, we want 29 days ago
            date_chicago = current_time_chicago - timedelta(days=i)
            date_start_chicago = date_chicago.replace(hour=0, minute=0, second=0, microsecond=0)
            date_end_chicago = date_start_chicago + timedelta(days=1)
            
            # Convert to UTC for database queries - handle timezone conversion properly
            if date_start_chicago.tzinfo is None:
                date_start_utc = chicago_tz.localize(date_start_chicago).astimezone(utc_tz)
            else:
                date_start_utc = date_start_chicago.astimezone(utc_tz)
                
            if date_end_chicago.tzinfo is None:
                date_end_utc = chicago_tz.localize(date_end_chicago).astimezone(utc_tz)
            else:
                date_end_utc = date_end_chicago.astimezone(utc_tz)
            
            touch_count = Log.query.filter(
                Log.created_at >= date_start_utc,
                Log.created_at < date_end_utc,
                Log.is_touch.is_(True)
            ).count()
            
            detailed_count = Log.query.filter(
                Log.created_at >= date_start_utc,
                Log.created_at < date_end_utc,
                Log.is_touch.is_(False)
            ).count()
            
            log_activity_data.append({
                'date': date_start_chicago.strftime('%Y-%m-%d'),
                'touch_logs': touch_count,
                'detailed_logs': detailed_count,
                'total_logs': touch_count + detailed_count
            })
    
    # Most logged projects (last 30 days)
    most_logged_projects = db.session.query(
        Project.name,
        Client.name.label('client_name'),
        func.count(Log.id).label('log_count'),
        func.sum(Log.hours).label('total_hours'),
        func.sum(text('CASE WHEN is_touch IS TRUE THEN 1 ELSE 0 END')).label('touch_count'),
        func.sum(text('CASE WHEN is_touch IS FALSE THEN 1 ELSE 0 END')).label('detailed_count')
    ).join(Client).join(Log).filter(
        Log.created_at >= thirty_days_ago_utc
    ).group_by(Project.id, Project.name, Client.name).order_by(desc('log_count')).limit(8).all()
    

    
    return render_template('analytics.html',
                         # Top metrics
                         total_members=total_members,
                         total_projects=total_projects,
                         total_clients=total_clients,
                         open_tasks=open_tasks,

                         # Task analytics
                         completion_data=completion_data,
                         most_active_projects=most_active_projects,
                         most_logged_time_projects=most_logged_time_projects,
                         # Log analytics
                         total_logs_last_30=total_logs_last_30,
                         touch_logs_last_30=touch_logs_last_30,
                         detailed_logs_last_30=detailed_logs_last_30,
                         total_hours_last_30=round(total_hours_last_30, 1),
                         log_activity_data=log_activity_data)

# LOGS ROUTES
@app.route('/logs')
def logs():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get all logs with their related data, ordered by most recent first
    logs = Log.query.join(
        User, User.id == Log.user_id
    ).outerjoin(
        Project, Project.id == Log.project_id
    ).outerjoin(
        Client, Client.id == Project.client_id
    ).order_by(Log.created_at.desc()).all()
    
    # Get projects and users for the edit form
    projects = Project.query.join(Client).filter(
        Project.status != 'Archived'
    ).order_by(Project.name.asc()).all()
    
    users = User.query.filter_by(role='admin').order_by(User.first_name.asc()).all()
    
    return render_template('logs.html', 
                         logs=logs,
                         projects=projects,
                         users=users)

@app.route('/edit_log/<int:log_id>', methods=['POST'])
def edit_log(log_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    log = Log.query.get_or_404(log_id)
    
    notes = request.form.get('notes', '').strip()
    hours = request.form.get('hours')
    fixed_cost = request.form.get('fixed_cost')
    project_id = request.form.get('project_id')
    log_datetime_str = request.form.get('log_datetime', '').strip()
    
    # Update log fields
    log.notes = notes if notes else None
    log.hours = float(hours) if hours else None
    log.fixed_cost = float(fixed_cost) if fixed_cost else None
    log.project_id = int(project_id) if project_id else None
    
    # Parse custom datetime if provided
    if log_datetime_str:
        try:
            from datetime import datetime
            log_datetime = datetime.strptime(log_datetime_str, '%Y-%m-%dT%H:%M')
            log_datetime = TIMEZONE.localize(log_datetime)
            log.created_at = log_datetime
        except ValueError:
            flash('Invalid date/time format', 'error')
            return redirect(url_for('logs'))
    
    db.session.commit()
    flash('Log entry updated successfully', 'success')
    return redirect(url_for('logs'))

@app.route('/delete_log/<int:log_id>', methods=['POST'])
def delete_log(log_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    log = Log.query.get_or_404(log_id)
    
    db.session.delete(log)
    db.session.commit()
    flash('Log entry deleted successfully', 'success')
    return redirect(url_for('logs'))

# LOGGING ROUTES
@app.route('/add_touch_log', methods=['POST'])
def add_touch_log():
    if 'user_id' not in session:
        return {'success': False, 'error': 'Not logged in'}, 401
    
    project_id = request.json.get('project_id')
    
    if not project_id:
        return {'success': False, 'error': 'Project ID required'}, 400
    
    # Verify project exists (allow any project including archived)
    project = Project.query.filter(Project.id == int(project_id)).first()
    if not project:
        return {'success': False, 'error': 'Project not found'}, 404
    
    # Create touch log with 20 minutes (0.333 hours)
    log = Log(
        is_touch=True,
        hours=0.333,  # 20 minutes
        user_id=session['user_id'],
        project_id=int(project_id)
    )
    
    db.session.add(log)
    db.session.commit()
    
    # Log touch activity
    ActivityLog.log_activity(
        user_id=session['user_id'],
        activity_type='time_logged',
        entity_type='log',
        entity_id=log.id,
        new_value={
            'is_touch': log.is_touch,
            'hours': log.hours,
            'project_id': log.project_id
        }
    )
    
    db.session.commit()
    
    return {'success': True, 'message': f'Touch logged for {project.name}'}

@app.route('/add_time_log', methods=['POST'])
def add_time_log():
    if 'user_id' not in session:
        return {'success': False, 'error': 'Not logged in'}, 401
    
    project_id = request.form.get('project_id')
    hours = request.form.get('hours')
    fixed_cost = request.form.get('fixed_cost')
    notes = request.form.get('notes', '').strip()
    log_datetime_str = request.form.get('log_datetime', '').strip()
    
    if not project_id:
        flash('Project is required', 'error')
        return redirect(request.referrer or url_for('dashboard'))
    
    if not notes:
        flash('Notes are required for time logs', 'error')
        return redirect(request.referrer or url_for('dashboard'))
    
    # Verify project exists (allow any project including archived)
    project = Project.query.filter(Project.id == int(project_id)).first()
    if not project:
        flash('Project not found', 'error')
        return redirect(request.referrer or url_for('dashboard'))
    
    # Parse custom datetime if provided
    log_datetime = None
    if log_datetime_str:
        try:
            # Parse the datetime-local format (YYYY-MM-DDTHH:MM)
            from datetime import datetime
            log_datetime = datetime.strptime(log_datetime_str, '%Y-%m-%dT%H:%M')
            # Localize to the application timezone
            log_datetime = TIMEZONE.localize(log_datetime)
        except ValueError:
            flash('Invalid date/time format', 'error')
            return redirect(request.referrer or url_for('dashboard'))
    
    # Create time log
    log = Log(
        is_touch=False,
        hours=float(hours) if hours else None,
        fixed_cost=float(fixed_cost) if fixed_cost else None,
        notes=notes,
        user_id=session['user_id'],
        project_id=int(project_id)
    )
    
    # Override created_at if custom datetime provided
    if log_datetime:
        log.created_at = log_datetime
    
    db.session.add(log)
    db.session.commit()
    
    # Log time logging activity with the same datetime
    activity_log = ActivityLog.log_activity(
        user_id=session['user_id'],
        activity_type='time_logged',
        entity_type='log',
        entity_id=log.id,
        new_value={
            'is_touch': log.is_touch,
            'hours': log.hours,
            'fixed_cost': float(log.fixed_cost) if log.fixed_cost else None,
            'project_id': log.project_id
        }
    )
    
    # Override ActivityLog created_at to match the log entry
    if log_datetime:
        activity_log.created_at = log_datetime
    
    db.session.commit()
    
    flash(f'Time logged for {project.name}', 'success')
    return redirect(request.referrer or url_for('dashboard'))

# PINNING AND FLAGGING ROUTES
@app.route('/project/<int:project_id>/pin', methods=['POST'])
def pin_project(project_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Check if already pinned
    existing_pin = UserProjectPin.query.filter_by(
        user_id=session['user_id'], 
        project_id=project_id
    ).first()
    
    if not existing_pin:
        pin = UserProjectPin(user_id=session['user_id'], project_id=project_id)
        db.session.add(pin)
        db.session.commit()
    
    return redirect(request.referrer or url_for('project_detail', project_id=project_id))

@app.route('/project/<int:project_id>/unpin', methods=['POST'])
def unpin_project(project_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    pin = UserProjectPin.query.filter_by(
        user_id=session['user_id'], 
        project_id=project_id
    ).first()
    
    if pin:
        db.session.delete(pin)
        db.session.commit()
    
    return redirect(request.referrer or url_for('project_detail', project_id=project_id))

@app.route('/task/<int:task_id>/flag', methods=['POST'])
def flag_task(task_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Check if already flagged
    existing_flag = UserTaskFlag.query.filter_by(
        user_id=session['user_id'], 
        task_id=task_id
    ).first()
    
    if not existing_flag:
        flag = UserTaskFlag(user_id=session['user_id'], task_id=task_id)
        db.session.add(flag)
        db.session.commit()
    
    return redirect(request.referrer or url_for('task_detail', task_id=task_id))

@app.route('/task/<int:task_id>/unflag', methods=['POST'])
def unflag_task(task_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    flag = UserTaskFlag.query.filter_by(
        user_id=session['user_id'], 
        task_id=task_id
    ).first()
    
    if flag:
        db.session.delete(flag)
        db.session.commit()
    
    return redirect(request.referrer or url_for('task_detail', task_id=task_id))

@app.route('/profile')
def profile():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = User.query.get_or_404(session['user_id'])
    
    # Get current date and time periods
    from datetime import datetime, timedelta
    import pytz
    
    central = pytz.timezone('America/Chicago')
    now = datetime.now(central)
    
    # Calculate time periods
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)
    
    # This week: Monday to Sunday
    days_since_monday = now.weekday()
    week_start = today_start - timedelta(days=days_since_monday)
    week_end = week_start + timedelta(days=7)
    
    # Last 30 days: 30 days ago to now
    thirty_days_ago = now - timedelta(days=30)
    thirty_days_start = thirty_days_ago.replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Convert to UTC for database queries
    today_start_utc = today_start.astimezone(pytz.UTC)
    today_end_utc = today_end.astimezone(pytz.UTC)
    week_start_utc = week_start.astimezone(pytz.UTC)
    week_end_utc = week_end.astimezone(pytz.UTC)
    thirty_days_start_utc = thirty_days_start.astimezone(pytz.UTC)
    now_utc = now.astimezone(pytz.UTC)
    
    # Calculate analytics for each time period
    analytics_data = {}
    
    for period_name, start_utc, end_utc in [
        ('today', today_start_utc, today_end_utc),
        ('this_week', week_start_utc, week_end_utc),
        ('last_30_days', thirty_days_start_utc, now_utc)
    ]:
        # Hours logged
        hours_logged = db.session.query(db.func.coalesce(db.func.sum(Log.hours), 0)).filter(
            Log.user_id == user.id,
            Log.created_at >= start_utc,
            Log.created_at < end_utc,
            Log.hours.isnot(None)
        ).scalar() or 0
        
        # Tasks created
        tasks_created = Task.query.filter(
            Task.created_by == user.id,
            Task.created_at >= start_utc,
            Task.created_at < end_utc
        ).count()
        
        # Tasks completed
        tasks_completed = Task.query.filter(
            Task.completed_by_user_id == user.id,
            Task.completed_on >= start_utc,
            Task.completed_on < end_utc
        ).count()
        
        # Projects worked on (unique projects from logs and tasks)
        # Get project IDs from logs
        log_project_ids = db.session.query(Log.project_id).filter(
            Log.user_id == user.id,
            Log.created_at >= start_utc,
            Log.created_at < end_utc,
            Log.project_id.isnot(None)
        ).distinct().all()
        
        # Get project IDs from tasks
        task_project_ids = db.session.query(Task.project_id).filter(
            db.or_(
                db.and_(Task.created_by == user.id, Task.created_at >= start_utc, Task.created_at < end_utc),
                db.and_(Task.completed_by_user_id == user.id, Task.completed_on >= start_utc, Task.completed_on < end_utc)
            ),
            Task.project_id.isnot(None)
        ).distinct().all()
        
        # Combine and count unique project IDs
        all_project_ids = set()
        for row in log_project_ids:
            if row[0] is not None:
                all_project_ids.add(row[0])
        for row in task_project_ids:
            if row[0] is not None:
                all_project_ids.add(row[0])
        
        projects_worked_on = len(all_project_ids)
        
        analytics_data[period_name] = {
            'hours_logged': hours_logged,
            'tasks_created': tasks_created,
            'tasks_completed': tasks_completed,
            'projects_worked_on': projects_worked_on
        }
    
    # Calculate statistical averages for business days and weekends (last 30 days only)
    # Get logs from the last 30 days only
    thirty_days_ago_utc = thirty_days_start_utc
    
    # Calculate total hours for last 30 days
    total_hours = db.session.query(db.func.coalesce(db.func.sum(Log.hours), 0)).filter(
        Log.user_id == user.id,
        Log.hours.isnot(None),
        Log.created_at >= thirty_days_ago_utc
    ).scalar() or 0
    
    # Get all log entries for the user from last 30 days
    all_logs = Log.query.filter(
        Log.user_id == user.id,
        Log.hours.isnot(None),
        Log.created_at >= thirty_days_ago_utc
    ).all()
    
    # Group by date and separate business days from weekends
    business_daily_hours = {}
    weekend_daily_hours = {}
    business_days_count = 0
    weekend_days_count = 0
    
    for log in all_logs:
        # Convert UTC to Central time
        log_date_central = log.created_at.astimezone(central)
        log_date = log_date_central.date()
        weekday = log_date_central.weekday()  # Monday=0, Sunday=6
        
        if weekday < 5:  # Monday-Friday (business days)
            if log_date not in business_daily_hours:
                business_daily_hours[log_date] = 0
                business_days_count += 1
            business_daily_hours[log_date] += log.hours or 0
        else:  # Saturday-Sunday (weekends)
            if log_date not in weekend_daily_hours:
                weekend_daily_hours[log_date] = 0
                weekend_days_count += 1
            weekend_daily_hours[log_date] += log.hours or 0
    
    # Calculate business day averages
    business_hours_total = sum(business_daily_hours.values())
    avg_hours_per_day = business_hours_total / max(1, business_days_count)
    
    # Calculate weekend day averages
    weekend_hours_total = sum(weekend_daily_hours.values())
    avg_hours_weekend_day = weekend_hours_total / max(1, weekend_days_count)
    
    # Calculate weekly averages (for last 30 days = ~4.3 weeks)
    weeks_in_30_days = 30 / 7
    avg_hours_per_week = total_hours / max(1, weeks_in_30_days)
    
    # Calculate weekend weekly average (weekends only)
    avg_hours_weekend_week = weekend_hours_total / max(1, weeks_in_30_days)
    
    # Calculate standard deviations
    business_hours_list = [hours for hours in business_daily_hours.values() if hours > 0]
    weekend_hours_list = [hours for hours in weekend_daily_hours.values() if hours > 0]
    
    import statistics
    std_dev_daily = statistics.stdev(business_hours_list) if len(business_hours_list) > 1 else 0
    std_dev_weekly = std_dev_daily * 5  # 5 business days per week
    
    stats = {
        'avg_hours_per_day': avg_hours_per_day,
        'avg_hours_weekend_day': avg_hours_weekend_day,
        'avg_hours_per_week': avg_hours_per_week,
        'avg_hours_weekend_week': avg_hours_weekend_week,
        'std_dev_daily': std_dev_daily,
        'std_dev_weekly': std_dev_weekly
    }
    
    return render_template('profile.html', user=user, analytics=analytics_data, stats=stats)

@app.route('/edit_profile', methods=['POST'])
def edit_profile():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user = User.query.get_or_404(session['user_id'])
    
    first_name = request.form.get('first_name', '').strip()
    last_name = request.form.get('last_name', '').strip()
    
    # Validation: first name is required
    if not first_name:
        flash('First name is required', 'error')
        return redirect(url_for('profile'))
    
    user.first_name = first_name
    user.set_last_name(last_name)
    
    db.session.commit()
    
    # Update session with new full name
    session['user_name'] = user.full_name
    
    flash('Profile updated successfully', 'success')
    return redirect(url_for('profile'))

@app.route('/supplement/<int:membership_id>/add', methods=['POST'])
def add_supplement(membership_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    membership = Membership.query.get_or_404(membership_id)
    
    time = request.form.get('time', type=int)
    budget = request.form.get('budget', type=float)
    notes = request.form.get('notes', '').strip()
    
    if not time and not budget:
        flash('Please specify either time or budget for the supplement.', 'warning')
        return redirect(url_for('membership_detail', membership_id=membership_id))
    
    supplement = MembershipSupplement(
        membership_id=membership_id,
        time=time,
        budget=budget,
        notes=notes
    )
    db.session.add(supplement)
    db.session.commit()
    
    # Log membership supplement activity
    ActivityLog.log_activity(
        user_id=session['user_id'],
        activity_type='membership_supplement_added',
        entity_type='membership_supplement',
        entity_id=supplement.id,
        new_value={
            'membership_id': supplement.membership_id,
            'time': supplement.time,
            'budget': supplement.budget
        }
    )
    
    db.session.commit()
    
    flash('Supplement added successfully.', 'success')
    return redirect(url_for('membership_detail', membership_id=membership_id))

@app.route('/supplement/<int:supplement_id>/edit', methods=['POST'])
def edit_supplement(supplement_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    supplement = MembershipSupplement.query.get_or_404(supplement_id)
    
    time = request.form.get('time', type=int)
    budget = request.form.get('budget', type=float)
    notes = request.form.get('notes', '').strip()
    
    if not time and not budget:
        flash('Please specify either time or budget for the supplement.', 'warning')
        return redirect(url_for('membership_detail', membership_id=supplement.membership_id))
    
    supplement.time = time
    supplement.budget = budget
    supplement.notes = notes
    db.session.commit()
    
    flash('Supplement updated successfully.', 'success')
    return redirect(url_for('membership_detail', membership_id=supplement.membership_id))

@app.route('/supplement/<int:supplement_id>/delete', methods=['POST'])
def delete_supplement(supplement_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    supplement = MembershipSupplement.query.get_or_404(supplement_id)
    membership_id = supplement.membership_id
    
    db.session.delete(supplement)
    db.session.commit()
    
    flash('Supplement deleted successfully.', 'success')
    return redirect(url_for('membership_detail', membership_id=membership_id))

@app.route('/api/supplement/<int:supplement_id>')
def get_supplement(supplement_id):
    if 'user_id' not in session:
        return {'error': 'Not authenticated'}, 401
    
    supplement = MembershipSupplement.query.get_or_404(supplement_id)
    return {
        'id': supplement.id,
        'time': supplement.time,
        'budget': supplement.budget,
        'notes': supplement.notes
    }

# ADMIN ROUTES (Admin only)
@app.route('/admin')
def admin():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Check if user is admin
    if session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('dashboard'))
    
    users = User.query.order_by(User.first_name.asc()).all()
    equipment_list = Equipment.query.order_by(Equipment.name.asc()).all()
    project_count = Project.query.count()
    
    # Get scheduling settings (create default if doesn't exist)
    scheduling_settings = SchedulingSettings.query.first()
    if not scheduling_settings:
        scheduling_settings = SchedulingSettings()
        db.session.add(scheduling_settings)
        db.session.commit()
    
    # Get all upcoming appointments (no time limit)
    from datetime import datetime
    import pytz
    
    # Use Central Time for display, but UTC for database queries
    central = pytz.timezone('America/Chicago')
    now_central = datetime.now(central)
    
    # Convert to UTC for database query (since database stores in UTC)
    utc = pytz.timezone('UTC')
    now_utc = now_central.astimezone(utc)
    start_date_utc = now_utc.replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Debug: Print query parameters
    print(f"\nDEBUG Admin Appointments Query:")
    print(f"Start date (Central): {now_central.replace(hour=0, minute=0, second=0, microsecond=0)}")
    print(f"Start date (UTC): {start_date_utc}")
    print(f"No end date limit - showing all future appointments")
    
    # First get all appointments to check what exists
    all_appointments = EquipmentAppointment.query.all()
    print(f"\nDEBUG All Appointments ({len(all_appointments)}):")
    for appt in all_appointments:
        print(f"ID: {appt.id}, User: {appt.user_id}, Equipment: {appt.equipment_id}, "
              f"Start: {appt.start_time}, Status: {appt.status}")
    
    # Now run the filtered query - show all future appointments including cancelled
    upcoming_appointments = EquipmentAppointment.query.filter(
        EquipmentAppointment.start_time >= start_date_utc
    ).order_by(EquipmentAppointment.start_time.asc()).all()
    
    print(f"\nDEBUG Filtered Appointments ({len(upcoming_appointments)}):")
    for appt in upcoming_appointments:
        print(f"ID: {appt.id}, Start: {appt.start_time}, Status: {appt.status}")
    
    # Get global blocked dates (future only)
    from datetime import date
    today = date.today()
    blocked_dates = EquipmentBlockedDate.query.filter(
        EquipmentBlockedDate.blocked_date >= today
    ).order_by(EquipmentBlockedDate.blocked_date.asc()).all()
    
    # Get global operating hours
    operating_hours = EquipmentOperatingHours.query.order_by(EquipmentOperatingHours.day_of_week.asc()).all()
    
    # Convert operating hours to JavaScript-friendly format
    operating_hours_js = {}
    for hours in operating_hours:
        operating_hours_js[hours.day_of_week] = {
            'start_time': hours.start_time.strftime('%H:%M'),
            'end_time': hours.end_time.strftime('%H:%M')
        }
    
    return render_template('admin.html', 
                         users=users, 
                         equipment_list=equipment_list, 
                         project_count=project_count,
                         scheduling_settings=scheduling_settings,
                         upcoming_appointments=upcoming_appointments,
                         blocked_dates=blocked_dates,
                         operating_hours=operating_hours,
                         operating_hours_js=operating_hours_js)

@app.route('/import_labs_projects', methods=['POST'])
def import_labs_projects():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('dashboard'))
    
    if 'import_file' not in request.files:
        flash('No file uploaded', 'error')
        return redirect(url_for('admin'))
    
    file = request.files['import_file']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('admin'))
    
    if not file.filename.endswith(('.json', '.txt')):
        flash('Invalid file type. Please upload a JSON or TXT file.', 'error')
        return redirect(url_for('admin'))
    
    try:
        content = file.read()
        data = json.loads(content)
        
        for client_data in data.get('Clients', []):
            # Check if client already exists
            client = Client.query.filter_by(name=client_data['Name']).first()
            if not client:
                client = Client(name=client_data['Name'])
                db.session.add(client)
                db.session.commit()
            
            # Add projects
            for project_name in client_data.get('Projects', []):
                # Check if project already exists for this client
                project = Project.query.filter_by(name=project_name, client_id=client.id).first()
                if not project:
                    project = Project(
                        name=project_name,
                        client_id=client.id,
                        status='Archived',
                        project_lead_id=1
                    )
                    db.session.add(project)
            
        db.session.commit()
        flash('Import completed successfully', 'success')
        
    except json.JSONDecodeError:
        flash('Invalid JSON format in uploaded file', 'error')
    except Exception as e:
        flash(f'Error during import: {str(e)}', 'error')
    
    return redirect(url_for('admin'))

@app.route('/run_import_script', methods=['POST'])
def run_import_script():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        import import_labs
        flash('Import script executed successfully', 'success')
    except Exception as e:
        flash(f'Error running import script: {str(e)}', 'error')
    
    return redirect(url_for('admin'))

@app.route('/add_equipment', methods=['POST'])
def add_equipment():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('dashboard'))
    
    name = request.form.get('name', '').strip()
    description = request.form.get('description', '').strip()
    manual = request.form.get('manual', '').strip()
    
    if not name:
        flash('Equipment name is required', 'error')
        return redirect(url_for('admin'))
    
    equipment = Equipment(
        name=name,
        description=description if description else None,
        manual=manual if manual else None,
        is_schedulable=request.form.get('is_schedulable') == 'on'
    )
    
    db.session.add(equipment)
    db.session.commit()
    flash('Equipment added successfully', 'success')
    return redirect(url_for('admin'))

@app.route('/edit_equipment/<int:equipment_id>', methods=['POST'])
def edit_equipment(equipment_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('dashboard'))
    
    equipment = Equipment.query.get_or_404(equipment_id)
    
    name = request.form.get('name', '').strip()
    description = request.form.get('description', '').strip()
    manual = request.form.get('manual', '').strip()
    
    if not name:
        flash('Equipment name is required', 'error')
        return redirect(url_for('admin'))
    
    equipment.name = name
    equipment.description = description if description else None
    equipment.manual = manual if manual else None
    equipment.is_schedulable = request.form.get('is_schedulable') == 'on'
    
    db.session.commit()
    flash('Equipment updated successfully', 'success')
    return redirect(url_for('admin'))

@app.route('/delete_equipment/<int:equipment_id>', methods=['POST'])
def delete_equipment(equipment_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('dashboard'))
    
    equipment = Equipment.query.get_or_404(equipment_id)
    
    db.session.delete(equipment)
    db.session.commit()
    flash('Equipment deleted successfully', 'success')
    return redirect(url_for('admin'))

@app.route('/add_user', methods=['POST'])
def add_user():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Check if user is admin
    if session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('dashboard'))
    
    first_name = request.form.get('first_name', '').strip()
    last_name = request.form.get('last_name', '').strip()
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '').strip()
    role = request.form.get('role', 'trainee')  # Default to trainee if not specified
    equipment_ids = request.form.getlist('equipment[]')
    
    if not first_name or not email:
        flash('First name and email are required', 'error')
        return redirect(url_for('admin'))
    
    # Check if email already exists
    existing_user = User.query.filter_by(email=email).first()
    if existing_user:
        flash('Email already exists', 'error')
        return redirect(url_for('admin'))
    
    # Validate role
    if role not in ['admin', 'trainee', 'finance']:
        flash('Invalid role specified', 'error')
        return redirect(url_for('admin'))
    
    # Only require password for admin users
    if role == 'admin' and not password:
        flash('Password is required for admin users', 'error')
        return redirect(url_for('admin'))
    
    user = User(
        first_name=first_name,
        email=email,
        role=role
    )
    user.set_last_name(last_name)
    if password:  # Only set password if provided
        user.set_password(password)
    
    # Add equipment associations
    if equipment_ids:
        # Convert string IDs to integers
        equipment_int_ids = [int(id) for id in equipment_ids if id.isdigit()]
        equipment_list = Equipment.query.filter(Equipment.id.in_(equipment_int_ids)).all()
        user.equipment.extend(equipment_list)
    
    db.session.add(user)
    db.session.commit()
    
    # Log user creation activity
    ActivityLog.log_activity(
        user_id=session['user_id'],
        activity_type='user_created',
        entity_type='user',
        entity_id=user.id,
        new_value={
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'role': user.role
        }
    )
    
    flash('User created successfully', 'success')
    return redirect(url_for('admin'))

@app.route('/edit_user/<int:user_id>', methods=['POST'])
def edit_user(user_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Check if user is admin
    if session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('dashboard'))
    
    user = User.query.get_or_404(user_id)
    
    first_name = request.form.get('first_name', '').strip()
    last_name = request.form.get('last_name', '').strip()
    email = request.form.get('email', '').strip()
    password = request.form.get('password', '').strip()
    role = request.form.get('role', 'trainee')  # Default to trainee if not specified
    equipment_ids = request.form.getlist('equipment[]')
    
    if not first_name or not email:
        flash('First name and email are required', 'error')
        return redirect(url_for('admin'))
    
    # Check if email already exists (excluding current user)
    existing_user = User.query.filter(User.email == email, User.id != user_id).first()
    if existing_user:
        flash('Email already exists', 'error')
        return redirect(url_for('admin'))
    
    # Validate role
    if role not in ['admin', 'trainee', 'finance']:
        flash('Invalid role specified', 'error')
        return redirect(url_for('admin'))
    
    # Only require password for admin users if changing to admin role
    if role == 'admin' and user.role != 'admin' and not password:
        flash('Password is required when promoting to admin', 'error')
        return redirect(url_for('admin'))
    
    user.first_name = first_name
    user.set_last_name(last_name)
    user.email = email
    user.role = role
    
    # Only update password if provided
    if password:
        user.set_password(password)
    
    # Update equipment associations
    user.equipment = []  # Clear existing associations
    if equipment_ids:
        # Convert string IDs to integers
        equipment_int_ids = [int(id) for id in equipment_ids if id.isdigit()]
        equipment_list = Equipment.query.filter(Equipment.id.in_(equipment_int_ids)).all()
        user.equipment.extend(equipment_list)
    
    db.session.commit()
    
    # Update session if editing own account
    if user_id == session['user_id']:
        session['user_name'] = user.full_name
        session['role'] = user.role
    
    flash('User updated successfully', 'success')
    return redirect(url_for('admin'))

@app.route('/delete_user/<int:user_id>', methods=['POST'])
def delete_user(user_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Check if user is admin
    if session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('dashboard'))
    
    # Prevent deleting yourself
    if user_id == session['user_id']:
        flash('Cannot delete your own account', 'error')
        return redirect(url_for('admin'))
    
    user = User.query.get_or_404(user_id)
    
    # Check if user has created tasks, logs, or leads projects
    has_data = (
        user.created_tasks.count() > 0 or
        user.assigned_tasks.count() > 0 or
        user.completed_tasks.count() > 0 or
        user.logs.count() > 0 or
        user.led_projects.count() > 0
    )
    
    if has_data:
        flash('Cannot delete user with existing data (tasks, logs, or led projects)', 'error')
        return redirect(url_for('admin'))
    
    db.session.delete(user)
    db.session.commit()
    flash('User deleted successfully', 'success')
    return redirect(url_for('admin'))



@app.route('/test-tasks')
def test_tasks():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get just a few tasks for testing
    test_tasks = Task.query.limit(3).all()
    
    # Serialize with minimal data
    tasks_data = []
    for task in test_tasks:
        tasks_data.append({
            'id': task.id,
            'description': task.description,
            'is_complete': task.is_complete,
            'project_name': task.project.name if task.project else None,
            'client_name': task.project.client.name if task.project and task.project.client else None,
            'creator_name': task.creator.full_name if task.creator else None,
            'assigned_to_id': task.assigned_to,
            'created_at': task.created_at.isoformat() if task.created_at else None
        })
    
    return render_template('test_tasks.html', tasks=tasks_data, current_user_id=session['user_id'])

@app.route('/api/preferences', methods=['GET'])
def get_preferences():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    preferences = UserPreferences.query.filter_by(user_id=session['user_id']).all()
    prefs_dict = {pref.key: pref.value for pref in preferences}
    
    return jsonify(prefs_dict)

@app.route('/api/preferences', methods=['POST'])
def save_preferences():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    data = request.get_json()
    if not data or not isinstance(data, dict):
        return jsonify({'error': 'Invalid data'}), 400
    
    # Update or create each preference
    for key, value in data.items():
        # Find existing preference
        pref = UserPreferences.query.filter_by(
            user_id=session['user_id'],
            key=key
        ).first()
        
        if pref:
            # Update existing
            pref.value = value
        else:
            # Create new
            pref = UserPreferences(
                user_id=session['user_id'],
                key=key,
                value=value
            )
            db.session.add(pref)
    
    db.session.commit()
    return jsonify({'status': 'success'})

@app.route('/api/task-form-data')
def get_task_form_data():
    if 'user_id' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    
    # Get all active projects
    projects = Project.query.join(
        Client, Client.id == Project.client_id
    ).filter(
        Project.status == 'Active'
    ).order_by(Project.name).all()
    
    # Get all users
    users = User.query.filter_by(role='admin').order_by(User.first_name).all()
    
    # Convert to dictionaries
    projects_data = [{
        'id': p.id,
        'name': p.name,
        'display_name': f"{p.name} ({p.client.name})" if p.client else p.name,
        'client_name': p.client.name if p.client else None,
        'is_default': p.is_default
    } for p in projects]
    
    users_data = [{
        'id': u.id,
        'name': u.full_name
    } for u in users]
    
    return jsonify({
        'projects': projects_data,
        'users': users_data
    })

@app.route('/api/projects_for_logging')
def get_projects_for_logging():
    if 'user_id' not in session:
        return {'projects': []}, 401
    
    # 1. Get the default project (if any)
    default_project = Project.query.filter(Project.is_default==True, Project.status!='Archived').first()
    default_project_id = default_project.id if default_project else None

    # 2. Get projects with most recent log entries (excluding default project)
    from sqlalchemy import desc
    recent_log_projects = db.session.query(
        Project, Client.name.label('client_name'), db.func.max(Log.created_at).label('last_log')
    ).join(Client).join(Log).filter(
        Project.status != 'Archived'
    )
    if default_project_id:
        recent_log_projects = recent_log_projects.filter(Project.id != default_project_id)
    recent_log_projects = recent_log_projects.group_by(Project.id, Client.name).order_by(desc('last_log')).all()

    # 3. Get the rest of the projects (excluding those already included)
    recent_log_project_ids = {p.id for p, _, _ in recent_log_projects}
    if default_project_id:
        recent_log_project_ids.add(default_project_id)
    other_projects = db.session.query(Project, Client.name.label('client_name')).join(Client).filter(
        Project.status != 'Archived',
        ~Project.id.in_(recent_log_project_ids)
    ).order_by(Project.updated_at.desc()).all()

    # 4. Get archived projects (not included above)
    all_nonarchived_ids = set(recent_log_project_ids)
    all_nonarchived_ids.update([p.id for p, _ in other_projects])
    archived_projects = db.session.query(Project, Client.name.label('client_name')).join(Client).filter(
        Project.status == 'Archived',
        ~Project.id.in_(all_nonarchived_ids)
    ).order_by(Project.name.asc()).all()

    # Build the ordered list
    projects = []
    if default_project:
        projects.append({
            'id': default_project.id,
            'name': default_project.name,
            'client_name': default_project.client.name if default_project.client else '',
            'display_name': f"{default_project.name} ({default_project.client.name})" if default_project.client else default_project.name,
            'status': default_project.status
        })
    for project, client_name, _ in recent_log_projects:
        projects.append({
            'id': project.id,
            'name': project.name,
            'client_name': client_name,
            'display_name': f"{project.name} ({client_name})",
            'status': project.status
        })
    for project, client_name in other_projects:
        projects.append({
            'id': project.id,
            'name': project.name,
            'client_name': client_name,
            'display_name': f"{project.name} ({client_name})",
            'status': project.status
        })
    for project, client_name in archived_projects:
        projects.append({
            'id': project.id,
            'name': project.name,
            'client_name': client_name,
            'display_name': f"{project.name} ({client_name})",
            'status': project.status
        })
    return {'projects': projects}

# Equipment Scheduling Routes

@app.route('/admin/scheduling_settings', methods=['POST'])
def update_scheduling_settings():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        max_duration = float(request.form.get('max_booking_duration_hours', 4.0))
        min_notice = float(request.form.get('min_booking_notice_hours', 4.0))
        advance_limit = int(request.form.get('booking_advance_limit_days', 7))
        
        settings = SchedulingSettings.query.first()
        if not settings:
            settings = SchedulingSettings()
            db.session.add(settings)
        
        settings.max_booking_duration_hours = max_duration
        settings.min_booking_notice_hours = min_notice
        settings.booking_advance_limit_days = advance_limit
        
        db.session.commit()
        flash('Scheduling settings updated successfully', 'success')
        
    except ValueError:
        flash('Invalid input values', 'error')
    except Exception as e:
        flash(f'Error updating settings: {str(e)}', 'error')
    
    return redirect(url_for('admin'))

@app.route('/admin/operating_hours', methods=['POST'])
def update_operating_hours():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('admin'))
    
    try:
        # Clear existing operating hours
        EquipmentOperatingHours.query.delete()
        
        # Add new operating hours
        for day in range(7):  # 0=Monday, 6=Sunday
            start_time_str = request.form.get(f'start_time_{day}')
            end_time_str = request.form.get(f'end_time_{day}')
            
            if start_time_str and end_time_str:
                from datetime import time
                start_time = time.fromisoformat(start_time_str)
                end_time = time.fromisoformat(end_time_str)
                
                operating_hours = EquipmentOperatingHours(
                    day_of_week=day,
                    start_time=start_time,
                    end_time=end_time
                )
                db.session.add(operating_hours)
        
        db.session.commit()
        flash('Operating hours updated successfully', 'success')
        
    except Exception as e:
        flash(f'Error updating operating hours: {str(e)}', 'error')
    
    return redirect(url_for('admin'))

@app.route('/admin/blocked_dates', methods=['POST'])
def add_blocked_date():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('admin'))
    
    try:
        blocked_date_str = request.form.get('blocked_date')
        reason = request.form.get('reason', '').strip()
        is_annual_recurring = request.form.get('is_annual_recurring') == 'on'
        
        if not blocked_date_str:
            flash('Blocked date is required', 'error')
            return redirect(url_for('admin'))
        
        from datetime import datetime
        blocked_date = datetime.strptime(blocked_date_str, '%Y-%m-%d').date()
        
        # Check if date already exists
        existing = EquipmentBlockedDate.query.filter_by(blocked_date=blocked_date).first()
        
        if existing:
            flash('This date is already blocked', 'error')
            return redirect(url_for('admin'))
        
        blocked_date_record = EquipmentBlockedDate(
            blocked_date=blocked_date,
            reason=reason,
            is_annual_recurring=is_annual_recurring
        )
        
        db.session.add(blocked_date_record)
        db.session.commit()
        
        flash('Blocked date added successfully', 'success')
        
    except ValueError:
        flash('Invalid date format', 'error')
    except Exception as e:
        flash(f'Error adding blocked date: {str(e)}', 'error')
    
    return redirect(url_for('admin'))

@app.route('/admin/blocked_date/<int:blocked_date_id>/delete', methods=['POST'])
def delete_blocked_date(blocked_date_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('admin'))
    
    blocked_date = EquipmentBlockedDate.query.get_or_404(blocked_date_id)
    
    try:
        db.session.delete(blocked_date)
        db.session.commit()
        flash('Blocked date removed successfully', 'success')
        
    except Exception as e:
        flash(f'Error removing blocked date: {str(e)}', 'error')
    
    return redirect(url_for('admin'))

@app.route('/admin/appointments/add', methods=['POST'])
def add_appointment():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('admin'))
    
    try:
        equipment_id = request.form.get('equipment_id')
        user_id = request.form.get('user_id')
        appointment_date = request.form.get('appointment_date')
        start_time_str = request.form.get('start_time')
        end_time_str = request.form.get('end_time')
        purpose = request.form.get('purpose', '').strip()
        notes = request.form.get('notes', '').strip()
        
        if not all([equipment_id, user_id, appointment_date, start_time_str, end_time_str]):
            flash('All required fields must be filled', 'error')
            return redirect(url_for('admin'))
        
        # Convert string IDs to integers
        try:
            equipment_id = int(equipment_id)
            user_id = int(user_id)
        except ValueError:
            flash('Invalid equipment or user ID', 'error')
            return redirect(url_for('admin'))
        
        from datetime import datetime
        # Combine date and time strings and localize to Chicago timezone
        central = pytz.timezone('America/Chicago')
        start_datetime_str = f"{appointment_date}T{start_time_str}:00"
        end_datetime_str = f"{appointment_date}T{end_time_str}:00"
        
        start_time = central.localize(datetime.fromisoformat(start_datetime_str))
        end_time = central.localize(datetime.fromisoformat(end_datetime_str))
        
        # Check if equipment is schedulable
        equipment = Equipment.query.get_or_404(equipment_id)
        if not equipment.is_schedulable:
            flash('This equipment is not available for scheduling', 'error')
            return redirect(url_for('admin'))
        
        # Check for conflicts
        conflicting_appointment = EquipmentAppointment.query.filter(
            EquipmentAppointment.equipment_id == equipment_id,
            EquipmentAppointment.status.in_(['approved', 'pending']),
            db.or_(
                db.and_(EquipmentAppointment.start_time <= start_time, EquipmentAppointment.end_time > start_time),
                db.and_(EquipmentAppointment.start_time < end_time, EquipmentAppointment.end_time >= end_time),
                db.and_(EquipmentAppointment.start_time >= start_time, EquipmentAppointment.end_time <= end_time)
            )
        ).first()
        
        if conflicting_appointment:
            flash('This time slot conflicts with an existing appointment', 'error')
            return redirect(url_for('admin'))
        
        # For admins, only check operating hours (no duration limits)
        # Get operating hours for the day of the week
        day_of_week = start_time.weekday()  # 0=Monday, 6=Sunday
        operating_hours = EquipmentOperatingHours.query.filter_by(day_of_week=day_of_week).first()
        
        if operating_hours:
            # Check if appointment is within operating hours
            start_time_only = start_time.time()
            end_time_only = end_time.time()
            
            if start_time_only < operating_hours.start_time or end_time_only > operating_hours.end_time:
                flash(f'Appointment must be within operating hours ({operating_hours.start_time.strftime("%H:%M")} - {operating_hours.end_time.strftime("%H:%M")})', 'error')
                return redirect(url_for('admin'))
        else:
            # No operating hours set for this day
            flash('No operating hours set for this day', 'error')
            return redirect(url_for('admin'))
        
        appointment = EquipmentAppointment(
            equipment_id=equipment_id,
            user_id=user_id,
            start_time=start_time,
            end_time=end_time,
            purpose=purpose,
            notes=notes,
            status='approved'  # Admin-created appointments start as approved
        )
        
        db.session.add(appointment)
        db.session.commit()
        
        flash('Appointment created successfully', 'success')
        
    except ValueError:
        flash('Invalid date/time format', 'error')
    except Exception as e:
        flash(f'Error creating appointment: {str(e)}', 'error')
    
    return redirect(url_for('admin'))

@app.route('/admin/appointment/<int:appointment_id>/status', methods=['POST'])
def update_appointment_status(appointment_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('admin'))
    
    appointment = EquipmentAppointment.query.get_or_404(appointment_id)
    new_status = request.form.get('status')
    
    if new_status not in ['pending', 'approved', 'cancelled']:
        flash('Invalid status', 'error')
        return redirect(url_for('admin'))
    
    try:
        old_status = appointment.status
        appointment.status = new_status
        db.session.commit()
        
        # Log the activity
        ActivityLog.log_activity(
            user_id=session['user_id'],
            activity_type='appointment_status_changed',
            entity_type='equipment_appointment',
            entity_id=appointment_id,
            old_value={'status': old_status},
            new_value={'status': new_status}
        )
        
        flash(f'Appointment status updated to {new_status}', 'success')
        
    except Exception as e:
        flash(f'Error updating appointment status: {str(e)}', 'error')
    
    return redirect(url_for('admin'))

@app.route('/schedule')
def schedule():
    """Public scheduling interface - step-by-step appointment booking"""
    return render_template('schedule.html')

@app.route('/schedule/<int:appointment_id>')
def appointment_detail(appointment_id):
    """Show appointment details and allow cancellation"""
    appointment = EquipmentAppointment.query.get_or_404(appointment_id)
    return render_template('appointment_detail.html', appointment=appointment)

@app.route('/schedule/<int:appointment_id>/calendar')
def appointment_calendar(appointment_id):
    """Generate and return an iCal file for the appointment"""
    appointment = EquipmentAppointment.query.get_or_404(appointment_id)
    
    # Create calendar
    cal = Calendar()
    cal.add('prodid', '-//Hub Tracker//Equipment Booking//EN')
    cal.add('version', '2.0')
    
    # Create event
    event = Event()
    event.add('summary', f'Equipment Booking: {appointment.equipment.name}')
    event.add('dtstart', appointment.start_time)
    event.add('dtend', appointment.end_time)
    event.add('dtstamp', appointment.created_at)
    
    # Add description with equipment details and purpose
    description = f"Equipment: {appointment.equipment.name}\n"
    if appointment.equipment.description:
        description += f"Details: {appointment.equipment.description}\n"
    if appointment.purpose:
        description += f"\nPurpose: {appointment.purpose}"
    event.add('description', description)
    
    # Add location if equipment has a manual URL (could be used to link to manual)
    if appointment.equipment.manual:
        event.add('url', appointment.equipment.manual)
    
    cal.add_component(event)
    
    # Save to bytes
    ical_data = cal.to_ical()
    
    # Create response
    response = send_file(
        BytesIO(ical_data),
        mimetype='text/calendar',
        as_attachment=True,
        download_name=f'equipment-booking-{appointment.id}.ics'
    )
    
    return response

@app.route('/api/validate_email', methods=['POST'])
def validate_email():
    """Validate user email and return user info if found"""
    data = request.get_json()
    email = data.get('email', '').strip().lower()
    
    if not email:
        return jsonify({'valid': False, 'message': 'Email is required'})
    
    user = User.query.filter_by(email=email).first()
    
    if not user:
        return jsonify({'valid': False, 'message': 'Email not found. Please contact an administrator.'})

    
    # Get all user's equipment regardless of schedulable status
    equipment_list = []
    for equipment in user.equipment.all():
        equipment_data = {
            'id': equipment.id,
            'name': equipment.name,
            'description': equipment.description or '',
            'manual': equipment.manual or '',
            'is_schedulable': equipment.is_schedulable,
            'contact_email': os.environ.get('CONTACT_EMAIL', 'neurotechhub@wustl.edu')  # Added contact email from env var
        }
        equipment_list.append(equipment_data)
    
    response_data = {
        'valid': True,
        'user': {
            'id': user.id,
            'name': user.full_name,
            'email': user.email
        },
        'equipment': equipment_list
    }
    
    return jsonify(response_data)

@app.route('/api/available_slots', methods=['POST'])
def get_available_slots():
    """Get available time slots for a specific equipment and date"""
    data = request.get_json()
    equipment_id = data.get('equipment_id')
    date_str = data.get('date')
    start_time = data.get('start_time')  # If provided, return possible end times
    
    if not equipment_id or not date_str:
        return jsonify({'error': 'Equipment ID and date are required'})
    
    try:
        # Get current time in America/Chicago timezone
        central = pytz.timezone('America/Chicago')
        now = datetime.now(central)
        
        # Parse date
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Get equipment
        equipment = Equipment.query.get_or_404(equipment_id)
        
        # Get scheduling settings
        settings = SchedulingSettings.query.first()
        if not settings:
            return jsonify({'error': 'Scheduling settings not configured'})
        
        # Get operating hours for this day
        day_of_week = selected_date.weekday()  # Monday=0, Sunday=6
        operating_hours = EquipmentOperatingHours.query.filter_by(day_of_week=day_of_week).first()
        
        if not operating_hours:
            return jsonify({'slots': [], 'message': 'Closed on this day'})
        
        # Get existing appointments for this equipment and date
        start_of_day = datetime.combine(selected_date, datetime.min.time())
        end_of_day = datetime.combine(selected_date, datetime.max.time())
        
        existing_appointments = EquipmentAppointment.query.filter(
            EquipmentAppointment.equipment_id == equipment_id,
            EquipmentAppointment.start_time >= start_of_day,
            EquipmentAppointment.start_time < end_of_day,
            EquipmentAppointment.status != 'cancelled'
        ).all()

        # If start_time is provided, return possible end times
        if start_time:
            start_dt = datetime.strptime(f"{date_str} {start_time}", '%Y-%m-%d %H:%M')
            start_dt = central.localize(start_dt)
            end_dt = datetime.combine(selected_date, operating_hours.end_time)
            end_dt = central.localize(end_dt)
            max_end_dt = start_dt + timedelta(hours=settings.max_booking_duration_hours)
            
            # Use the earlier of max duration or operating hours end
            end_dt = min(end_dt, max_end_dt)
            
            slots = []
            current_dt = start_dt + timedelta(minutes=30)  # Start with minimum duration (30 min)
            
            while current_dt <= end_dt:
                # Check if slot conflicts with existing appointments
                is_available = True
                for appointment in existing_appointments:
                    if (start_dt < appointment.end_time and current_dt > appointment.start_time):
                        is_available = False
                        break
                
                if is_available:
                    slots.append({
                        'end_time': current_dt.strftime('%H:%M'),
                        'formatted': f"Until {current_dt.strftime('%I:%M %p')}"
                    })
                
                current_dt += timedelta(minutes=30)
            
            return jsonify({
                'slots': slots,
                'type': 'end_times'
            })
        
        # Otherwise, return available start times
        slots = []
        current_dt = datetime.combine(selected_date, operating_hours.start_time)
        current_dt = central.localize(current_dt)
        end_dt = datetime.combine(selected_date, operating_hours.end_time)
        end_dt = central.localize(end_dt)
        
        # If today, start from current time rounded up to next 30-minute increment
        if selected_date == now.date() and now > current_dt:
            current_dt = now.replace(second=0, microsecond=0)
            minutes = current_dt.minute
            if minutes % 30 != 0:
                minutes = ((minutes // 30) + 1) * 30
                if minutes == 60:
                    current_dt = current_dt + timedelta(hours=1)
                    current_dt = current_dt.replace(minute=0)
                else:
                    current_dt = current_dt.replace(minute=minutes)
        
        # Don't allow start times that would exceed max duration before end of operating hours
        latest_start = end_dt - timedelta(minutes=30)  # Minimum 30 min duration
        
        while current_dt <= latest_start:
            # Check if slot conflicts with existing appointments
            is_available = True
            for appointment in existing_appointments:
                if (current_dt < appointment.end_time and 
                    current_dt + timedelta(minutes=30) > appointment.start_time):
                    is_available = False
                    break
            
            if is_available:
                slots.append({
                    'start_time': current_dt.strftime('%H:%M'),
                    'formatted': current_dt.strftime('%I:%M %p')
                })
            
            current_dt += timedelta(minutes=30)
        
        return jsonify({
            'slots': slots,
            'type': 'start_times'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/available_dates', methods=['POST'])
def get_available_dates():
    """Get available dates for a specific equipment"""
    data = request.get_json()
    equipment_id = data.get('equipment_id')
    
    if not equipment_id:
        return jsonify({'error': 'Equipment ID is required'})
    
    try:
        # Get scheduling settings
        settings = SchedulingSettings.query.first()
        if not settings:
            return jsonify({'error': 'Scheduling settings not configured'})
        
        # Get equipment
        equipment = Equipment.query.get_or_404(equipment_id)
        
        # Get blocked dates
        today = date.today()
        blocked_dates = EquipmentBlockedDate.query.filter(
            EquipmentBlockedDate.blocked_date >= today
        ).all()
        blocked_date_set = {bd.blocked_date for bd in blocked_dates}
        
        # Get current time in America/Chicago timezone
        central = pytz.timezone('America/Chicago')
        now = datetime.now(central)
        current_date = now.date()
        
        # Generate available dates
        available_dates = []
        check_date = current_date
        
        for i in range(settings.booking_advance_limit_days + 1):
            # Check if date is blocked
            if check_date in blocked_date_set:
                check_date += timedelta(days=1)
                continue
            
            # Check if we have operating hours for this day
            day_of_week = check_date.weekday()
            operating_hours = EquipmentOperatingHours.query.filter_by(day_of_week=day_of_week).first()
            
            if operating_hours:
                # For today, check if there's still enough time left in operating hours
                if check_date == current_date:
                    # Convert operating hours end time to datetime
                    end_time = datetime.combine(check_date, operating_hours.end_time)
                    end_time = central.localize(end_time)
                    
                    # Need at least 30 minutes before end of operating hours
                    if (end_time - now).total_seconds() < 1800:  # 30 minutes in seconds
                        check_date += timedelta(days=1)
                        continue
                    
                    # If current time is past operating hours start time,
                    # check if there's at least one 30-minute slot available
                    start_time = datetime.combine(check_date, operating_hours.start_time)
                    start_time = central.localize(start_time)
                    if now > start_time:
                        # Round up to next 30-minute increment
                        minutes = now.minute
                        if minutes % 30 != 0:
                            minutes = ((minutes // 30) + 1) * 30
                            if minutes == 60:
                                now = now + timedelta(hours=1)
                                now = now.replace(minute=0)
                            else:
                                now = now.replace(minute=minutes)
                        
                        # If rounded time is past end time minus 30 minutes, skip this day
                        if (end_time - now).total_seconds() < 1800:
                            check_date += timedelta(days=1)
                            continue
                
                available_dates.append({
                    'date': check_date.strftime('%Y-%m-%d'),
                    'day_name': check_date.strftime('%A'),
                    'formatted': check_date.strftime('%A, %B %d, %Y')
                })
            
            check_date += timedelta(days=1)
        
        return jsonify({'dates': available_dates})
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/create_appointment', methods=['POST'])
def create_public_appointment():
    """Create appointment from public interface"""
    data = request.get_json()
    
    try:
        # Validate required fields
        required_fields = ['user_id', 'equipment_id', 'date', 'start_time', 'end_time']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} is required'}), 400
        
        # Parse date and times
        appointment_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        start_time = datetime.strptime(data['start_time'], '%H:%M').time()
        end_time = datetime.strptime(data['end_time'], '%H:%M').time()
        
        # Create appointment datetime with Chicago timezone
        central = pytz.timezone('America/Chicago')
        start_datetime = central.localize(datetime.combine(appointment_date, start_time))
        end_datetime = central.localize(datetime.combine(appointment_date, end_time))
        
        # Create appointment
        appointment = EquipmentAppointment(
            equipment_id=data['equipment_id'],
            user_id=data['user_id'],
            start_time=start_datetime,
            end_time=end_datetime,
            purpose=data.get('purpose', ''),
            notes=data.get('notes', ''),
            status='approved'  # Auto-approve for public bookings
        )
        
        db.session.add(appointment)
        db.session.commit()
        
        # Send email notification only for public appointments (not admin-created)
        try:
            send_appointment_notification(appointment)
        except Exception as e:
            print(f"Warning: Failed to send email notification: {e}")
            # Don't fail the appointment creation if email fails
        
        return jsonify({
            'success': True,
            'appointment_id': appointment.id,
            'redirect_url': url_for('appointment_detail', appointment_id=appointment.id)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/cancel_appointment/<int:appointment_id>', methods=['POST'])
def cancel_public_appointment(appointment_id):
    """Cancel appointment from public interface"""
    appointment = EquipmentAppointment.query.get_or_404(appointment_id)
    
    if appointment.status == 'cancelled':
        return jsonify({'error': 'Appointment is already cancelled'}), 400
    
    appointment.status = 'cancelled'
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Appointment cancelled successfully'})

@app.route('/admin/appointment/<int:appointment_id>/delete', methods=['POST'])
def delete_appointment(appointment_id):
    """Delete an appointment"""
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('admin'))
    
    appointment = EquipmentAppointment.query.get_or_404(appointment_id)
    
    try:
        db.session.delete(appointment)
        db.session.commit()
        flash('Appointment deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting appointment: {str(e)}', 'error')
    
    return redirect(url_for('admin'))

@app.route('/admin/appointment/<int:appointment_id>/edit', methods=['POST'])
def edit_appointment(appointment_id):
    """Edit an existing appointment"""
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Access denied. Administrator privileges required.', 'error')
        return redirect(url_for('dashboard'))
    
    appointment = EquipmentAppointment.query.get_or_404(appointment_id)
    
    # Get form data
    equipment_id = request.form.get('equipment_id')
    user_id = request.form.get('user_id')
    appointment_date = request.form.get('appointment_date')
    start_time = request.form.get('start_time')
    end_time = request.form.get('end_time')
    purpose = request.form.get('purpose')
    notes = request.form.get('notes')
    status = request.form.get('status')
    
    if not all([equipment_id, user_id, appointment_date, start_time, end_time, status]):
        flash('All required fields must be filled out.', 'error')
        return redirect(url_for('admin'))
    
    try:
        # Parse date and times
        date_obj = datetime.strptime(appointment_date, '%Y-%m-%d').date()
        start_dt = datetime.combine(date_obj, datetime.strptime(start_time, '%H:%M').time())
        end_dt = datetime.combine(date_obj, datetime.strptime(end_time, '%H:%M').time())
        
        # Add timezone info (Central Time)
        central = pytz.timezone('America/Chicago')
        start_dt = central.localize(start_dt)
        end_dt = central.localize(end_dt)
        
        # Update appointment
        appointment.equipment_id = equipment_id
        appointment.user_id = user_id
        appointment.start_time = start_dt
        appointment.end_time = end_dt
        appointment.purpose = purpose
        appointment.notes = notes
        appointment.status = status
        
        db.session.commit()
        flash('Appointment updated successfully.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating appointment: {str(e)}', 'error')
    
    return redirect(url_for('admin'))

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 5001))) 